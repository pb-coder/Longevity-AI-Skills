"""One-shot migrator: monthly xlsx sheets → per-month CSVs.

Reads every ``YYYY.MM`` sheet from ``Workout Tracker - <Person>.xlsx``,
writes each as ``<person>/data/monthly/YYYY.MM.csv`` (18-column schema
preserved verbatim), then archives the xlsx as
``Workout Tracker - <Person>.pre-monthly-csv-backup.xlsx``.

Idempotent: re-running on a clean target produces no-op writes
(``canonicalize_monthly_csv`` makes the on-disk shape deterministic).

This is the LAST place ``openpyxl`` should appear in the import graph.
After this script runs, every other consumer reads/writes CSV directly.

Usage:
    python3 migrate_xlsx_monthly_to_csv.py --person Nihad
    python3 migrate_xlsx_monthly_to_csv.py --person Nihad --dry-run
    python3 migrate_xlsx_monthly_to_csv.py --person Nihad --keep-xlsx
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from monthly_csv import (  # noqa: E402
    MONTHLY_COLS,
    MONTHLY_FIELDS,
    TOTAL_LABEL,
    canonicalize_monthly_csv,
    upsert_rows,
)
from person_paths import (  # noqa: E402
    ensure_monthly_dir,
    monthly_csv as monthly_csv_path,
    tracker_for,
)

MONTHLY_RE = re.compile(r"^\d{4}\.\d{2}$")


def _read_sheet_rows(ws) -> list[dict]:
    """Pull every populated row from a YYYY.MM sheet into the dict shape
    that ``monthly_csv.upsert_rows`` expects.

    Includes TOTAL rows so canonicalize can re-hoist Apple-watch session
    metadata into a fresh TOTAL row. Drops fully-empty rows.
    """
    out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        padded = list(row) + [None] * MONTHLY_COLS
        cells = padded[:MONTHLY_COLS]
        if all(c is None or c == "" for c in cells):
            continue
        d = {}
        for i, key in enumerate(MONTHLY_FIELDS):
            v = cells[i]
            if isinstance(v, str) and v.strip() == "":
                d[key] = None
            else:
                d[key] = v
        out.append(d)
    return out


def migrate(person: str, dry_run: bool = False, keep_xlsx: bool = False) -> int:
    """Read every ``YYYY.MM`` sheet from the per-person xlsx, write CSVs."""
    xlsx = tracker_for(person)
    if not xlsx.exists():
        print(f"ERROR: tracker xlsx not found: {xlsx}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    monthly_sheet_names = sorted([
        n for n in wb.sheetnames if MONTHLY_RE.match(n)
    ])
    if not monthly_sheet_names:
        print(f"ERROR: no monthly sheets in {xlsx}", file=sys.stderr)
        return 1

    summaries: list[str] = []
    for ym in monthly_sheet_names:
        ws = wb[ym]
        rows = _read_sheet_rows(ws)
        non_total = [r for r in rows if not (
            isinstance(r.get("exercise"), str)
            and r["exercise"].strip().upper() == TOTAL_LABEL
        )]
        if dry_run:
            summaries.append(f"{ym}: {len(non_total)} rows (dry-run, not written)")
            continue
        # Wipe the target so re-runs are deterministic.
        target = monthly_csv_path(person, ym)
        if target.exists():
            target.unlink()
        ensure_monthly_dir(person)
        upsert_rows(person, ym, rows)
        # Verify the written CSV has at least the data-row count from xlsx.
        from monthly_csv import read_monthly
        written = read_monthly(person, ym)
        written_non_total = [r for r in written if not (
            isinstance(r.get("exercise"), str)
            and r["exercise"].strip().upper() == TOTAL_LABEL
        )]
        diff = len(non_total) - len(written_non_total)
        marker = "" if diff == 0 else f" (Δ={diff} vs xlsx)"
        summaries.append(
            f"{ym}: {len(non_total)} rows → {target.name}{marker}"
        )

    if not dry_run and not keep_xlsx:
        backup = xlsx.with_suffix(".pre-monthly-csv-backup.xlsx")
        xlsx.rename(backup)
        summaries.append(f"Archived: {xlsx.name} → {backup.name}")
    elif keep_xlsx and not dry_run:
        summaries.append(f"--keep-xlsx: {xlsx.name} left in place")

    for line in summaries:
        print(line)
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--person", required=True,
                    help="Tracker owner (Nihad or Fabian).")
    ap.add_argument("--dry-run", action="store_true",
                    help="Read xlsx + report row counts; don't write CSVs.")
    ap.add_argument("--keep-xlsx", action="store_true",
                    help="Don't archive the xlsx after a successful migration.")
    args = ap.parse_args()
    return migrate(args.person, dry_run=args.dry_run, keep_xlsx=args.keep_xlsx)


if __name__ == "__main__":
    sys.exit(main())
