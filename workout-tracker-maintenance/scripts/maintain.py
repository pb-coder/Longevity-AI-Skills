"""Workout Tracker monthly maintenance.

Idempotent. Safe to re-run. Performs:
  1. Restyle the monthly YYYY.MM workout sheets to the canonical look
  2. Trim empty trailing rows/columns (with buffer for the current month)
  3. Reorder remaining sheets: months newest → oldest
  4. Validate the per-person CSVs (well-formed, monotonic dates)
  5. Report row counts and verify data integrity

Post-PR1: only the monthly workout sheets live in xlsx; Health Metrics,
Workout Sessions, and Profile are CSVs in ``<person>/data/``.

Usage:
    python3 maintain.py --person Nihad
    python3 maintain.py --person Nihad --dry-run
    python3 maintain.py --person Nihad --fix-distance-units
"""
import argparse
import csv
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "shared"))
from tracker_sheet import (  # noqa: E402
    MONTHLY_COLS,
    _format_pace_min_per_km,
    _parse_duration_minutes,
    canonicalize_sheet_order,
    date_str,
    find_last_data_cell,
    style_monthly_sheet,
)
from csv_store import (  # noqa: E402
    HEALTH_METRICS_HEADERS_BY_SOURCE,
    WORKOUT_SESSIONS_HEADERS_BY_SOURCE,
    read_profile,
)
from person_paths import (  # noqa: E402
    data_dir,
    health_metrics_csv,
    profile_csv,
    tracker_for,
    workout_sessions_csv,
)

# Row buffer policy when trimming empty rows.
# Current-month sheet: leave a generous buffer since the user is actively logging.
# Past-month sheets: 2 blank rows.
CURRENT_MONTH_BUFFER = 50
PAST_MONTH_BUFFER = 2


# ------------------------------------------------------------------ helpers
def is_monthly(sheet_name: str) -> bool:
    return bool(re.match(r"^\d{4}\.\d{2}$", sheet_name))


def current_month_key():
    """Return the YYYY.MM string for the current calendar month."""
    return datetime.now().strftime("%Y.%m")


# ------------------------------------------------------------------ trim
def trim_sheet(ws, buffer: int, target_cols: int):
    """Drop trailing empty rows (keeping `buffer` extras) and cap columns at `target_cols`."""
    last_r, last_c = find_last_data_cell(ws)
    if last_r < 1:
        last_r = 1
    target_last_row = last_r + buffer

    if ws.max_row > target_last_row:
        ws.delete_rows(target_last_row + 1, ws.max_row - target_last_row)
    if ws.max_column > target_cols:
        ws.delete_cols(target_cols + 1, ws.max_column - target_cols)


# ------------------------------------------------------------------ reorder
def reorder_sheets(wb):
    """Delegate to the shared canonicalize_sheet_order helper.

    Kept as a thin wrapper so existing callers (and tests, if any) don't
    need to update their call sites — the canonical order logic lives in
    ``tracker_sheet`` so writers (`/log`, importers, /maintain) all use
    the same code path.
    """
    canonicalize_sheet_order(wb)


# ------------------------------------------------------------------ main
def run(person: str, dry_run: bool = False) -> int:
    """Restyle the workout xlsx (monthly sheets only) and validate the CSVs."""
    path = tracker_for(person)
    if not path.exists():
        print(f"ERROR: tracker not found: {path}", file=sys.stderr)
        return 1

    # Safety backup before any write.
    if not dry_run:
        backup = path.with_suffix(".maintain-backup.xlsx")
        shutil.copy2(path, backup)
        print(f"Backup: {backup.name}")

    wb = openpyxl.load_workbook(path)

    # 0. Drop empty monthly sheets that aren't the current month — keeps the
    # workbook lean. Empty sheets get created by accident (e.g. a future-month
    # placeholder) and clutter the tab bar. The current month is preserved
    # even when empty because /log may write to it any moment.
    current = current_month_key()
    for name in list(wb.sheetnames):
        if is_monthly(name) and name != current:
            if count_nonempty_rows(wb[name]) == 0:
                del wb[name]
                print(f"Dropped empty sheet: {name}")

    # Warn if a stale dense sheet survived the migration. Past-PR1 the
    # workbook should hold only YYYY.MM monthly sheets.
    stale_sheets = [
        n for n in wb.sheetnames
        if not is_monthly(n)
    ]
    if stale_sheets:
        print(f"WARN: unexpected non-monthly sheets in xlsx: {stale_sheets} "
              f"(should have been migrated to CSV by migrate_xlsx_to_csv.py)")

    before_counts = {name: count_nonempty_rows(wb[name]) for name in wb.sheetnames}

    # 1. Style + trim every monthly sheet.
    for name in list(wb.sheetnames):
        if not is_monthly(name):
            continue
        ws = wb[name]
        style_monthly_sheet(ws)
        buf = CURRENT_MONTH_BUFFER if name == current else PAST_MONTH_BUFFER
        trim_sheet(ws, buffer=buf, target_cols=MONTHLY_COLS)

    # 2. Reorder
    reorder_sheets(wb)

    # 3. Verify data integrity (nonempty row counts should be preserved)
    after_counts = {name: count_nonempty_rows(wb[name]) for name in wb.sheetnames}
    mismatches = [n for n in before_counts if before_counts[n] != after_counts.get(n)]
    if mismatches:
        print(f"ERROR: nonempty row count changed: {mismatches}", file=sys.stderr)
        for n in mismatches:
            print(f"  {n}: before={before_counts[n]} after={after_counts.get(n)}")
        return 2

    # 4. Write
    if dry_run:
        print("Dry run — not saving.")
    else:
        wb.save(path)
        print(f"Saved: {path.name}")

    print("\nFinal sheet order:", list(wb.sheetnames))
    print("\nRow counts per sheet:")
    for name in wb.sheetnames:
        ws = wb[name]
        last_r, last_c = find_last_data_cell(ws)
        print(f"  {name}: data rows={after_counts[name]:4d}  max_row={ws.max_row:4d}  max_col={ws.max_column}")

    # 5. Validate CSVs (per-person data folder).
    csv_status = validate_csvs(person)
    print("\nCSV checks:")
    for line in csv_status:
        print(f"  {line}")
    return 0


def validate_csvs(person: str) -> list[str]:
    """Sanity-check the per-person CSVs.

    Reports header schema match (against the active source's expected
    header), monotonic-DESC date order, and row counts. Read-only — never
    rewrites the CSVs; the importers' upsert helpers own the rewrite path.
    """
    out: list[str] = []
    profile = read_profile(person)
    source = profile.get("source") or "xml"
    if source not in HEALTH_METRICS_HEADERS_BY_SOURCE:
        source = "xml"

    targets = {
        "health_metrics.csv":   (health_metrics_csv(person),
                                 HEALTH_METRICS_HEADERS_BY_SOURCE[source]),
        "workout_sessions.csv": (workout_sessions_csv(person),
                                 WORKOUT_SESSIONS_HEADERS_BY_SOURCE[source]),
        "profile.csv":          (profile_csv(person),
                                 ["key", "value"]),
    }

    for label, (path, expected_header) in targets.items():
        if not path.exists():
            out.append(f"{label}: missing")
            continue
        with path.open("r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            try:
                header = next(reader)
            except StopIteration:
                out.append(f"{label}: empty (no header)")
                continue
            rows = list(reader)
        if header != expected_header:
            out.append(
                f"{label}: header mismatch — got {header}, "
                f"expected {expected_header}"
            )
            continue
        # Monotonic-DESC date check (Health Metrics + Workout Sessions only).
        if label != "profile.csv":
            dates = [row[0] for row in rows if row and row[0]]
            if dates and any(dates[i] < dates[i + 1] for i in range(len(dates) - 1)):
                out.append(f"{label}: WARN dates not strictly DESC")
        out.append(f"{label}: {len(rows)} rows ok")
    return out


def count_nonempty_rows(ws) -> int:
    """Count non-empty rows.

    Monthly sheets: count rows with an Exercise value populated, excluding
    TOTAL summary rows. Exercise sits at col C pre-migration, col D post-
    migration; we read both and take whichever is populated. This ignores
    legacy trailing rows that only held a carried-down Volume formula —
    they have no exercise name, so they don't represent logged sets either
    before or after migration. Keeps the before/after verify stable.

    Other sheets: count any row with content.
    """
    if is_monthly(ws.title):
        count = 0
        for row in ws.iter_rows(min_row=2):
            ex_c = row[2].value if len(row) > 2 else None  # col C
            ex_d = row[3].value if len(row) > 3 else None  # col D
            exercise = ex_c if isinstance(ex_c, str) else ex_d
            if not isinstance(exercise, str) or not exercise.strip():
                continue
            if exercise.strip().upper() == "TOTAL":
                continue
            count += 1
        return count

    count = 0
    for row in ws.iter_rows():
        if any(c.value is not None and c.value != "" for c in row):
            count += 1
    return count


# ------------------------------------------------------------------ historical fix
# Distance threshold above which a Swim row's value is almost certainly a
# meter-as-km bug. Open-water records exceed 30 km but the tracker's user
# base is pool-only; >10 km is functionally never legitimate here.
SUSPICIOUS_SWIM_DISTANCE_KM = 10.0


def _is_swim(exercise) -> bool:
    return isinstance(exercise, str) and exercise.strip().lower() == "swim"


def _is_swim_session_type(apple_type) -> bool:
    return isinstance(apple_type, str) and "swimming" in apple_type.lower()


def fix_distance_units(person: str, dry_run: bool = False) -> int:
    """Scan all monthly + Workout Sessions rows for the meter-as-km bug.

    Auto-fix swim rows whose Distance (km) > 10 (almost certainly metres):
    divide by 1000, recompute pace via the shared formatter (which now
    blanks degenerate ``0:01`` outputs). Print every change. For other
    activities (Run / Cycle / Walk / Hike), only flag suspicious rows
    where pace is implausibly fast — never auto-mutate them, since the
    XML importer never had the unit bug for non-swims.

    Idempotent: re-runs on a clean tracker print "no fixes needed" and
    don't touch the file.
    """
    path = tracker_for(person)
    if not path.exists():
        print(f"ERROR: tracker not found: {path}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(path)
    fixes: list[tuple[str, int, str]] = []
    flags: list[tuple[str, int, str]] = []

    # 1. Monthly xlsx sweep: swims with distance > 10 km auto-fix; non-swim
    # outliers flagged but not mutated.
    for name in wb.sheetnames:
        if not is_monthly(name):
            continue
        ws = wb[name]
        for r in range(2, ws.max_row + 1):
            exercise = ws.cell(row=r, column=4).value
            distance_v = ws.cell(row=r, column=10).value
            duration_v = ws.cell(row=r, column=11).value
            if distance_v in (None, ""):
                continue
            try:
                distance = float(distance_v)
            except (TypeError, ValueError):
                continue

            if _is_swim(exercise) and distance > SUSPICIOUS_SWIM_DISTANCE_KM:
                new_distance = round(distance / 1000.0, 3)
                dur_min = _parse_duration_minutes(duration_v)
                new_pace = _format_pace_min_per_km(dur_min, new_distance)
                old_pace = ws.cell(row=r, column=12).value
                if not dry_run:
                    ws.cell(row=r, column=10).value = new_distance
                    ws.cell(row=r, column=12).value = new_pace
                fixes.append((
                    name, r,
                    f"Swim {date_str(ws.cell(row=r, column=2).value)}: "
                    f"distance {distance} → {new_distance} km, "
                    f"pace {old_pace!r} → {new_pace!r}"
                ))
                continue

            # Sanity-flag non-swim rows where pace is implausibly fast.
            dur_min = _parse_duration_minutes(duration_v)
            if dur_min and distance > 0:
                pace = dur_min / distance
                if pace < 0.5:
                    flags.append((
                        name, r,
                        f"{exercise} {date_str(ws.cell(row=r, column=2).value)}: "
                        f"pace {pace:.3f} min/km — verify distance/duration"
                    ))

    # 2. Workout Sessions CSV sweep — Apple-Watch swim rows where
    # ``Distance (km) > 10`` get the same divide-by-1000 fix.
    ws_path = workout_sessions_csv(person)
    if ws_path.exists():
        with ws_path.open("r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            header = next(reader, [])
            csv_rows = list(reader)
        if header:
            try:
                date_idx = header.index("Date")
                type_idx = header.index("Apple Type")
                dist_idx = header.index("Distance (km)")
            except ValueError:
                date_idx = type_idx = dist_idx = -1
            if date_idx >= 0 and type_idx >= 0 and dist_idx >= 0:
                csv_changed = False
                for i, row in enumerate(csv_rows):
                    if len(row) <= max(date_idx, type_idx, dist_idx):
                        continue
                    apple_type = row[type_idx]
                    if not _is_swim_session_type(apple_type):
                        continue
                    if not row[dist_idx]:
                        continue
                    try:
                        distance = float(row[dist_idx])
                    except ValueError:
                        continue
                    if distance > SUSPICIOUS_SWIM_DISTANCE_KM:
                        new_distance = round(distance / 1000.0, 3)
                        if not dry_run:
                            row[dist_idx] = str(new_distance)
                            csv_changed = True
                        fixes.append((
                            "workout_sessions.csv", i + 2,
                            f"Swimming {row[date_idx]}: "
                            f"distance {distance} → {new_distance} km"
                        ))
                if csv_changed and not dry_run:
                    with ws_path.open("w", encoding="utf-8", newline="") as f:
                        writer = csv.writer(f)
                        writer.writerow(header)
                        writer.writerows(csv_rows)

    if not fixes and not flags:
        print(f"{path.name}: no fixes needed")
        return 0

    print(f"{path.name}:")
    for sheet, row, msg in fixes:
        print(f"  fix [{sheet} row {row}]: {msg}")
    for sheet, row, msg in flags:
        print(f"  flag [{sheet} row {row}]: {msg}")

    # Re-style every monthly sheet that received a fix so the corrected
    # Pace cell is laid out properly. Cheap and idempotent.
    if fixes and not dry_run:
        touched_monthly = sorted({s for s, _r, _m in fixes if is_monthly(s)})
        for s in touched_monthly:
            style_monthly_sheet(wb[s])

    if dry_run:
        print(f"\nDry run — {len(fixes)} fixes would be applied, "
              f"{len(flags)} flags reviewed (no save).")
    else:
        if fixes:
            backup = path.with_suffix(".fix-units-backup.xlsx")
            shutil.copy2(path, backup)
            print(f"\nBackup: {backup.name}")
            wb.save(path)
            print(f"Saved: {path.name} ({len(fixes)} fixes applied, "
                  f"{len(flags)} flags reviewed)")
        else:
            print(f"\n{len(flags)} flags reviewed; no auto-fixes applied.")
    return 0


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--person", required=True,
                    help="Tracker owner (Nihad or Fabian).")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument(
        "--fix-distance-units", action="store_true",
        help="Run the meter-as-km historical sweep for swim rows",
    )
    args = ap.parse_args()
    if args.fix_distance_units:
        sys.exit(fix_distance_units(args.person, args.dry_run))
    sys.exit(run(args.person, args.dry_run))
