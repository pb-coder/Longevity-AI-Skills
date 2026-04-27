"""canonicalize_logs.py — Rename typo'd exercise names in past monthly sheets and
strip stale '(not in database)' notes for rows whose exercise is now canonical.

Iterates every YYYY.MM sheet in a tracker xlsx, applies the rename map (case-
insensitive) to the Exercise column, and removes '(not in database)' from the
Notes column whenever the (post-rename) exercise name is in the canonical
exercises-database.md. After all edits, calls style_monthly_sheet so SESSION
numbering, sort, merges, and TOTAL rows self-heal.

Ambiguous names (e.g. bare 'Leg Curl' which could be Lying or Seated) are
reported but not auto-renamed.

Usage:
    python3 canonicalize_logs.py "Workout Tracker - Nihad.xlsx"
"""
from __future__ import annotations
import re
import sys
from pathlib import Path

import openpyxl

THIS_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(THIS_DIR))
from tracker_sheet import style_monthly_sheet  # noqa: E402

# --- canonical-name source --------------------------------------------------

DB_MD = THIS_DIR / "exercises-database.md"


def load_canonical_names(md_path: Path) -> set[str]:
    """Lowercased set of canonical exercise names from exercises-database.md."""
    names: set[str] = set()
    for raw in md_path.read_text().splitlines():
        s = raw.rstrip()
        if not s.startswith("- "):
            continue
        body = s[2:].strip()
        if body.startswith("(") or ":" in body.split("[", 1)[0]:
            continue
        if "[" in body:
            head = body.split("[", 1)[0]
        elif "—" in body:
            head = body.split("—", 1)[0]
        else:
            head = body
        head = head.strip()
        if head:
            names.add(head.lower())
    return names


# --- rename rules -----------------------------------------------------------

RENAMES: dict[str, str] = {
    "deadhang": "Dead Hang",
    "dead hang": "Dead Hang",
    "dips": "Dip",
    "triceps pushdown": "Cable Tricep Pushdown",
    "tricep pushdown": "Cable Tricep Pushdown",
    "scapular pull ups": "Scapular Pull-Up",
    "scapular pullups": "Scapular Pull-Up",
    "hanging leg raise": "Leg Raise",
    "lying leg curl": "Leg Curl (Lying)",
    "stomach crunch": "Ab Crunch Machine",
    "stomach crunches": "Ab Crunch Machine",
    "crunch": "Ab Crunch Machine",
    "stomach press": "Ab Crunch Machine",
    "stomach press vertical": "Ab Crunch Machine",
    "stomach press vertical machine": "Ab Crunch Machine",
}

# Keys we report (don't auto-rename) so the user can disambiguate.
AMBIGUOUS = {"leg curl"}

# Some renames also need a Notes annotation (e.g. Hanging Leg Raise → Leg Raise
# with a 'hanging' note so the variant info isn't lost).
RENAME_NOTE: dict[str, str] = {
    "hanging leg raise": "hanging",
}

NOT_IN_DB_RE = re.compile(r"\s*\(\s*not in database\s*\)\s*", re.IGNORECASE)
SHEET_NAME_RE = re.compile(r"^\d{4}\.\d{2}$")

EXERCISE_COL = 4
NOTES_COL = 9


# --- core --------------------------------------------------------------------

def canonicalize_sheet(ws, canonical: set[str]) -> tuple[int, int, list[tuple[int, str]]]:
    """Apply renames and clear stale notes on a single sheet.

    Returns (renamed_count, cleared_notes_count, ambiguous_rows).
    ambiguous_rows is [(row, original_name)] for caller-side reporting.
    """
    renamed = 0
    cleared = 0
    ambiguous: list[tuple[int, str]] = []

    for r in range(2, ws.max_row + 1):
        ex_cell = ws.cell(row=r, column=EXERCISE_COL)
        ex_val = ex_cell.value
        if not ex_val or not isinstance(ex_val, str):
            continue
        key = ex_val.strip().lower()

        if key in AMBIGUOUS:
            ambiguous.append((r, ex_val))

        if key in RENAMES:
            ex_cell.value = RENAMES[key]
            renamed += 1
            note_addition = RENAME_NOTE.get(key)
            if note_addition:
                notes_cell = ws.cell(row=r, column=NOTES_COL)
                cur = (notes_cell.value or "").strip()
                if note_addition not in cur.lower():
                    notes_cell.value = note_addition if not cur else f"{cur}; {note_addition}"

        # After potential rename, check whether the canonical name is in the DB
        # and the Notes column carries '(not in database)' that's now stale.
        post_name = (ex_cell.value or "").strip().lower()
        if post_name in canonical:
            notes_cell = ws.cell(row=r, column=NOTES_COL)
            notes_val = notes_cell.value
            if isinstance(notes_val, str) and NOT_IN_DB_RE.search(notes_val):
                cleaned = NOT_IN_DB_RE.sub("", notes_val).strip().strip(";").strip()
                notes_cell.value = cleaned or None
                cleared += 1

    return renamed, cleared, ambiguous


def main(xlsx_path: str) -> None:
    canonical = load_canonical_names(DB_MD)
    wb = openpyxl.load_workbook(xlsx_path)

    monthly_sheets = [s for s in wb.sheetnames if SHEET_NAME_RE.match(s)]
    if not monthly_sheets:
        print(f"{xlsx_path}: no monthly sheets found")
        return

    total_renamed = 0
    total_cleared = 0
    all_ambiguous: list[tuple[str, int, str]] = []  # (sheet, row, name)

    for name in monthly_sheets:
        ws = wb[name]
        renamed, cleared, amb = canonicalize_sheet(ws, canonical)
        if renamed or cleared or amb:
            print(f"  {name}: renamed={renamed} cleared_notes={cleared} ambiguous={len(amb)}")
        total_renamed += renamed
        total_cleared += cleared
        for r, nm in amb:
            all_ambiguous.append((name, r, nm))

        # Re-style the sheet so SESSION/sort/merges/TOTAL rebuild around the edits.
        # The styler rebuilds the data area from scratch by reading the cells we
        # just modified, so renamed names propagate cleanly.
        if renamed or cleared:
            style_monthly_sheet(ws)

    wb.save(xlsx_path)
    print(f"{xlsx_path}: total renamed={total_renamed} cleared_notes={total_cleared}")

    if all_ambiguous:
        print("\nAmbiguous rows (manual disambiguation needed):")
        print(f"  {'Sheet':<10} {'Row':>4}  Exercise (decide Lying vs Seated)")
        for sheet, row, nm in all_ambiguous:
            print(f"  {sheet:<10} {row:>4}  {nm}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("usage: canonicalize_logs.py <tracker.xlsx>", file=sys.stderr)
        sys.exit(2)
    main(sys.argv[1])
