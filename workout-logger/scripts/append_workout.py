"""Append parsed workout rows to the right monthly sheet in the given tracker xlsx.

Routes each row to the YYYY.MM sheet matching its date. Creates the sheet
(headers only) if missing. Styling is applied on every write via the shared
`tracker_sheet.style_monthly_sheet` so new rows match the rest of the sheet
without waiting for /maintain.

Input JSON is either a bare list of row dicts (legacy) or a wrapper object:

    {
      "rows": [ ... row dicts ... ],
      "bodyweight": [ {"date": "YYYY-MM-DD", "kg": 78.4, "notes": ""}, ... ]
    }

The wrapper form allows /log to capture the user's morning weight alongside
the workout. Both `rows` and `bodyweight` are optional within the wrapper.
Bodyweight entries are upserted into the per-person Health Metrics CSV
(``<person>/data/health_metrics.csv`` col ``Bodyweight (kg)``).

Row dict schema:
    {
      "date": "2026-04-20",          # required, YYYY-MM-DD
      "num": 1,                      # exercise index within the session
      "exercise": "Dumbbell Flat Bench Press",
      "set": 1,
      "reps": 10,
      "kg": 52,
      "volume": 520,
      "notes": "",
      "distance_km": null,           # optional cardio fields
      "duration_min": null,
      "pace": null,                  # string "MM:SS"
      "avg_hr": null,
      "laps": null                   # swim only
    }

Rows must arrive pre-sorted: by date ascending, then by num ascending, then by set.
The script does not re-sort — it trusts the caller.

Usage:
    python3 append_workout.py --person Nihad <payload_json_path>
    python3 append_workout.py --person Nihad -    # read JSON from stdin
"""
import argparse
import json
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "shared"))
from tracker_sheet import (  # noqa: E402
    MONTHLY_HEADERS,
    TOTAL_LABEL,
    _numeric_cell,
    canonicalize_sheet_order,
    style_monthly_sheet,
)
from csv_store import upsert_health_metrics  # noqa: E402
from person_paths import tracker_for  # noqa: E402

HEADERS = MONTHLY_HEADERS


def sheet_for_date(date_str: str) -> str:
    """'2026-04-20' -> '2026.04'."""
    y, m, _ = date_str.split("-")
    return f"{y}.{m}"


def find_last_data_row(ws) -> int:
    """Return the last row containing real set data.

    Trailing TOTAL rows are skipped so new sets land at the session boundary,
    not after the TOTAL. When `style_monthly_sheet` runs after the append it
    rebuilds TOTAL rows in the correct place.
    """
    last = 1
    for row in ws.iter_rows(min_row=2):
        # Column D (index 3) is Exercise in the 13-col layout.
        exercise = row[3].value if len(row) > 3 else None
        if exercise == TOTAL_LABEL:
            continue
        if any(c.value is not None and c.value != "" for c in row):
            last = row[0].row
    return last


def ensure_sheet(wb, name: str):
    """Return the sheet, creating it with headers if missing."""
    if name in wb.sheetnames:
        return wb[name], False
    ws = wb.create_sheet(title=name)
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)
    return ws, True


def row_values(r: dict):
    """Build the 18-col row. SESSION and Volume are left blank — the styler
    populates the SESSION number (merged per date) and writes the Volume
    formula (``=F*G``) on every set row. Date/#/Exercise are populated on
    every row (tracker convention).

    Numeric columns are coerced to real ints/floats via ``_numeric_cell`` so
    Excel can consume them in arithmetic formulas — stringified numbers like
    ``"67,5"`` would silently break ``=F*G`` and produce #VALUE!. The
    styler also normalises on every write, but pre-coercing here means any
    intermediate inspection of the sheet shows clean types from the start.
    """
    return [
        None,  # SESSION — styler fills & merges
        r["date"],
        _numeric_cell(r["num"]),
        r["exercise"],
        _numeric_cell(r["set"]),
        _numeric_cell(r.get("reps")),
        _numeric_cell(r.get("kg")),
        None,  # Volume — styler writes =F*G formula
        r.get("notes") or None,
        _numeric_cell(r.get("distance_km")),
        r.get("duration_min"),  # MM:SS string — not arithmetic
        r.get("pace"),           # MM:SS string — not arithmetic
        _numeric_cell(r.get("avg_hr")),
        # Cols 14-17: Apple-Watch session metadata. Manual /log payloads
        # don't supply these — the importers fill them on the first row
        # of the matching session via upsert_monthly_strength_session /
        # upsert_monthly_cardio.
        None,  # Active Cal
        None,  # Total Cal
        None,  # Elevation (m)
        None,  # Elapsed
        # Col 18: swim Laps. Filled by the Apple importer from the XML
        # lap-event count, or by the user via /log when they include
        # `<N> laps` / `<N> lengths` / `<N> bahnen` on a swim row.
        _numeric_cell(r.get("laps")),
    ]


def upsert_bodyweight(person: str, entries: list[dict]) -> list[str]:
    """Mirror manual bodyweight entries into the Health Metrics CSV.

    Bodyweight is no longer a separate sheet — it lives on the
    ``Bodyweight (kg)`` column of ``<person>/data/health_metrics.csv``.
    Each ``{"date": ..., "kg": ..., "notes": ...}`` becomes a Health
    Metrics record with ``bodyweight_kg`` set; csv_store's sparse-merge
    leaves all other metrics on that date alone.
    """
    if not entries:
        return []
    metric_entries = []
    for e in entries:
        d = str(e["date"])[:10]
        try:
            kg = float(e["kg"])
        except (TypeError, ValueError):
            continue
        metric_entries.append({"date": d, "bodyweight_kg": kg})
    if not metric_entries:
        return []
    upsert_health_metrics(person, metric_entries)
    summary = ", ".join(f"{e['date']}={e['kg']}kg" for e in entries)
    return [f"Bodyweight: mirrored to Health Metrics ({summary})"]


def write_payload(person: str, rows: list[dict], bodyweight: list[dict]) -> list[str]:
    """Apply rows + bodyweight entries in a single save."""
    tracker_path = tracker_for(person)
    if not tracker_path.exists():
        raise FileNotFoundError(f"tracker xlsx not found: {tracker_path}")
    wb = openpyxl.load_workbook(tracker_path)
    status = []

    if rows:
        by_sheet: dict[str, list[dict]] = {}
        for r in rows:
            by_sheet.setdefault(sheet_for_date(r["date"]), []).append(r)
        for sheet_name, sheet_rows in by_sheet.items():
            ws, created = ensure_sheet(wb, sheet_name)
            # Legacy 12-col sheets had ``Date`` in column A; the current
            # layout is ``SESSION | Date | …``. Trigger a full restyle only
            # when we detect the old header — otherwise rely on the post-
            # write restyle to handle sort/merge/TOTAL rebuilds. Skipping
            # the redundant pre-pass roughly halves ``/log`` run time on
            # large sheets.
            if ws.cell(row=1, column=1).value == "Date":
                style_monthly_sheet(ws)
            last_row = find_last_data_row(ws)
            write_row = last_row + 1
            for r in sheet_rows:
                for col, val in enumerate(row_values(r), start=1):
                    ws.cell(row=write_row, column=col, value=val)
                write_row += 1
            style_monthly_sheet(ws)
            dates = sorted({r["date"] for r in sheet_rows})
            tag = " (new sheet)" if created else ""
            status.append(
                f"Appended {len(sheet_rows)} row(s) to {sheet_name}{tag} "
                f"for {', '.join(dates)}"
            )

    # Re-canonicalize sheet order so any newly-created month sheet
    # lands in the right tab position without waiting for /maintain.
    canonicalize_sheet_order(wb)
    wb.save(tracker_path)

    # Bodyweight upserts the Health Metrics CSV (separate file from the xlsx).
    status.extend(upsert_bodyweight(person, bodyweight))
    return status


def load_payload(source: str) -> tuple[list[dict], list[dict]]:
    """Return (rows, bodyweight_entries). Accepts bare list or wrapper dict."""
    if source == "-":
        data = json.load(sys.stdin)
    else:
        data = json.loads(Path(source).read_text())

    if isinstance(data, list):
        rows = data
        bw = []
    elif isinstance(data, dict):
        rows = data.get("rows", []) or []
        bw_raw = data.get("bodyweight", []) or []
        # Accept a single dict as shorthand for a one-element list.
        if isinstance(bw_raw, dict):
            bw_raw = [bw_raw]
        bw = bw_raw
    else:
        raise ValueError("payload must be a list of rows or a dict wrapper")

    for r in rows:
        if "date" not in r or "exercise" not in r or "set" not in r or "num" not in r:
            raise ValueError(f"row missing required field: {r!r}")
    for e in bw:
        if "date" not in e or "kg" not in e:
            raise ValueError(f"bodyweight entry missing date/kg: {e!r}")

    return rows, bw


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--person", required=True,
                    help="Tracker owner (Nihad or Fabian).")
    ap.add_argument("payload", type=str,
                    help="Path to payload JSON, or '-' to read from stdin.")
    args = ap.parse_args()

    rows, bodyweight = load_payload(args.payload)
    if not rows and not bodyweight:
        print("No rows or bodyweight entries to write.")
        return 0
    try:
        for line in write_payload(args.person, rows, bodyweight):
            print(line)
    except FileNotFoundError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
