"""One-shot migration: dense xlsx sheets → per-person CSVs.

Steps for a single person:

1. Resolve the existing tracker xlsx (handles legacy
   ``./Workout Tracker - <Person>.xlsx`` at the root and the post-
   migration ``<Person>/Workout Tracker - <Person>.xlsx``).
2. Move the xlsx into ``<root>/<Person>/`` if it's still at the root.
3. For each of ``Health Metrics``, ``Workout Sessions``, ``Profile``,
   read the xlsx sheet → write the corresponding CSV in
   ``<Person>/data/``. Refuses to clobber a non-empty CSV unless
   ``--force`` is passed.
4. Drop the four dense sheets (``Exercises Database``, ``Profile``,
   ``Health Metrics``, ``Workout Sessions``) from the xlsx; save.
5. Print a summary: row counts written, sheets dropped, current xlsx
   tab list.

Idempotent. Re-running with no source sheets to migrate prints
"already migrated" and changes nothing. A pre-migration backup is
written next to the xlsx with the suffix ``.pre-csv-backup.xlsx``.

Usage:
    python3 Skills/shared/migrate_xlsx_to_csv.py --person Nihad
    python3 Skills/shared/migrate_xlsx_to_csv.py --person Nihad --dry-run
    python3 Skills/shared/migrate_xlsx_to_csv.py --person Nihad --force
"""
from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from csv_store import (  # noqa: E402
    HEALTH_METRICS_FIELDS_BY_SOURCE,
    HEALTH_METRICS_HEADERS_BY_SOURCE,
    PROFILE_KEYS,
    WORKOUT_SESSIONS_FIELDS_BY_SOURCE,
    WORKOUT_SESSIONS_HEADERS_BY_SOURCE,
    upsert_health_metrics,
    upsert_workout_sessions,
    write_profile,
)
from person_paths import (  # noqa: E402
    ensure_data_dir,
    health_metrics_csv,
    legacy_root_tracker_for,
    person_dir,
    profile_csv,
    tracker_for,
    workout_sessions_csv,
)

DROP_SHEETS = (
    "Exercises Database",
    "Profile",
    "Health Metrics",
    "Workout Sessions",
    # Legacy: a Bodyweight tab existed pre-2026-05 before being merged
    # into Health Metrics. Drop it if any tracker still has one.
    "Bodyweight",
)


def _resolve_xlsx(person: str, dry_run: bool) -> Path | None:
    """Find the tracker xlsx — prefer the post-migration location.

    Returns None if neither the new nor the legacy path resolves. In
    dry-run we report what would be moved but do not move; in the live
    run we move the legacy file into the per-person folder.
    """
    new = tracker_for(person)
    if new.exists():
        return new
    legacy = legacy_root_tracker_for(person)
    if not legacy.exists():
        return None
    if dry_run:
        print(f"[dry-run] would move {legacy.name} → {new}")
        return legacy
    person_dir(person).mkdir(parents=True, exist_ok=True)
    legacy.rename(new)
    print(f"Moved: {legacy.name} → {new.relative_to(legacy.parent)}")
    return new


def _read_health_metrics_from_xlsx(wb, source: str) -> list[dict]:
    """Translate the xlsx Health Metrics sheet into csv_store entry dicts."""
    if "Health Metrics" not in wb.sheetnames:
        return []
    ws = wb["Health Metrics"]
    fields = HEALTH_METRICS_FIELDS_BY_SOURCE[source]
    headers = HEALTH_METRICS_HEADERS_BY_SOURCE[source]
    out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        d = str(row[0])[:10] if row[0] is not None else None
        if not d or len(d) != 10:
            continue
        rec = {"date": d}
        for i, key in enumerate(fields, start=1):
            v = row[i] if len(row) > i else None
            rec[key] = v
        notes = row[len(headers) - 1] if len(row) >= len(headers) else None
        if notes not in (None, ""):
            rec["notes"] = notes
        out.append(rec)
    return out


def _read_workout_sessions_from_xlsx(wb, source: str) -> list[dict]:
    """Translate the xlsx Workout Sessions sheet into csv_store entry dicts."""
    if "Workout Sessions" not in wb.sheetnames:
        return []
    ws = wb["Workout Sessions"]
    fields = WORKOUT_SESSIONS_FIELDS_BY_SOURCE[source]
    out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] in (None, ""):
            continue
        d = str(row[0])[:10]
        if len(d) != 10:
            continue
        rec = {"date": d}
        for i, key in enumerate(fields, start=1):
            v = row[i] if len(row) > i else None
            rec[key] = v
        out.append(rec)
    return out


def _read_profile_from_xlsx(wb) -> dict:
    """Read the xlsx Profile sheet into a {key: value} dict."""
    if "Profile" not in wb.sheetnames:
        return {}
    ws = wb["Profile"]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        k = (row[0] or "").strip().lower() if row[0] else None
        v = row[1] if len(row) > 1 else None
        if k in PROFILE_KEYS:
            out[k] = v
    return out


def migrate(person: str, dry_run: bool, force: bool) -> int:
    """Run the migration. Return process exit code."""
    xlsx = _resolve_xlsx(person, dry_run)
    if xlsx is None:
        print(f"ERROR: no tracker xlsx found for {person}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(xlsx)

    # Decide which dense sheets need migration vs are already gone.
    todo_sheets = [s for s in DROP_SHEETS if s in wb.sheetnames]
    if not todo_sheets:
        print(f"{person}: already migrated (no dense sheets present)")
        return 0

    # CSV-overwrite safety: refuse to clobber non-empty existing CSVs.
    targets = {
        "Health Metrics":   health_metrics_csv(person),
        "Workout Sessions": workout_sessions_csv(person),
        "Profile":          profile_csv(person),
    }
    for sheet_name, target_path in targets.items():
        if sheet_name not in todo_sheets:
            continue
        if target_path.exists() and target_path.stat().st_size > 0 and not force:
            print(
                f"ERROR: {target_path.relative_to(person_dir(person))} already exists "
                f"and is non-empty. Pass --force to overwrite, or move the file aside "
                f"first.",
                file=sys.stderr,
            )
            return 2

    # Determine source from xlsx Profile (defaults to xml).
    profile_dict = _read_profile_from_xlsx(wb)
    source = profile_dict.get("source") or "xml"
    if source not in ("xml", "hl_export"):
        source = "xml"

    summary = {}

    # Read dense sheets while xlsx still has them.
    hm_entries = _read_health_metrics_from_xlsx(wb, source)
    ws_entries = _read_workout_sessions_from_xlsx(wb, source)

    if dry_run:
        print(f"[dry-run] {person}: would migrate from {xlsx.name}")
        if "Health Metrics" in todo_sheets:
            print(f"  health_metrics.csv: {len(hm_entries)} rows")
        if "Workout Sessions" in todo_sheets:
            print(f"  workout_sessions.csv: {len(ws_entries)} rows")
        if "Profile" in todo_sheets:
            print(f"  profile.csv: {len(profile_dict)} keys → {profile_dict}")
        print(f"  drop sheets: {todo_sheets}")
        print(f"  remaining xlsx sheets: "
              f"{[s for s in wb.sheetnames if s not in todo_sheets]}")
        return 0

    # Backup before mutating.
    backup = xlsx.with_suffix(".pre-csv-backup.xlsx")
    shutil.copy2(xlsx, backup)
    print(f"Backup: {backup.name}")

    ensure_data_dir(person)

    # Force-overwrite path: clear the target CSVs so the upsert starts
    # fresh rather than sparse-merging into stale data.
    if force:
        for target_path in targets.values():
            if target_path.exists():
                target_path.unlink()

    # Profile first: the upserts read it for source resolution.
    if "Profile" in todo_sheets:
        # write_profile fills defaults for any missing key.
        write_profile(person, **profile_dict)
        summary["profile.csv"] = f"{len(profile_dict)} keys"

    if "Health Metrics" in todo_sheets:
        upsert_health_metrics(person, hm_entries)
        summary["health_metrics.csv"] = f"{len(hm_entries)} rows"

    if "Workout Sessions" in todo_sheets:
        upsert_workout_sessions(person, ws_entries)
        summary["workout_sessions.csv"] = f"{len(ws_entries)} rows"

    # Drop the dense sheets from the xlsx and save.
    for s in todo_sheets:
        del wb[s]
    wb.save(xlsx)

    print(f"{person}: migrated {xlsx.name}")
    for k, v in summary.items():
        print(f"  wrote {k}: {v}")
    print(f"  dropped sheets: {sorted(todo_sheets)}")
    print(f"  remaining xlsx sheets: {wb.sheetnames}")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--person", required=True, help="Nihad or Fabian")
    ap.add_argument("--dry-run", action="store_true",
                    help="Report what would happen; don't write anything")
    ap.add_argument("--force", action="store_true",
                    help="Overwrite non-empty existing CSVs")
    args = ap.parse_args()
    return migrate(args.person, args.dry_run, args.force)


if __name__ == "__main__":
    sys.exit(main())