"""Check whether the Bodyweight sheet already has an entry for a given date.

Usage:
    python3 check_bodyweight.py "Workout Tracker.xlsx" YYYY-MM-DD

Prints a single-line JSON object:
    {"has_today": bool, "last": {"date": "YYYY-MM-DD", "kg": float} | null}

The `last` field is the most recent entry on or before the queried date,
useful for showing recent trend context when prompting the user. If the
Bodyweight sheet is missing entirely, returns has_today=false, last=null.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl


def iso_date(value) -> str | None:
    """Accept either a datetime/date object or an ISO string; return 'YYYY-MM-DD' or None."""
    if value is None or value == "":
        return None
    if hasattr(value, "isoformat"):
        s = value.isoformat()
        return s[:10]
    return str(value)[:10]


def _find_date_in_row(row: tuple) -> tuple[str | None, int | None]:
    """Return (YYYY-MM-DD, index) for the first date-shaped value in the first
    two columns — tolerating both the 4-col layout (Year|Date|Kg|Notes) and
    the legacy 3-col layout (Date|Kg|Notes)."""
    for i, v in enumerate(row[:2]):
        if v is None or v == "":
            continue
        s = iso_date(v)
        if s and len(s) == 10 and s[4] == "-" and s[7] == "-":
            return s, i
    return None, None


def check(tracker_path: Path, target_date: str) -> dict:
    if not tracker_path.exists():
        print(f"ERROR: tracker not found: {tracker_path}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(tracker_path, read_only=True, data_only=True)
    if "Bodyweight" not in wb.sheetnames:
        return {"has_today": False, "last": None}

    ws = wb["Bodyweight"]
    has_today = False
    best = None  # (date_str, kg) with date_str <= target_date, most recent

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        date_str, date_idx = _find_date_in_row(row)
        if date_str is None or date_idx is None:
            continue
        kg = row[date_idx + 1] if len(row) > date_idx + 1 else None
        if kg is None:
            continue
        if date_str == target_date:
            has_today = True
        if date_str <= target_date:
            if best is None or date_str > best[0]:
                best = (date_str, kg)

    last = None
    if best is not None:
        try:
            last = {"date": best[0], "kg": float(best[1])}
        except (TypeError, ValueError):
            last = None

    return {"has_today": has_today, "last": last}


def main() -> int:
    if len(sys.argv) != 3:
        print(__doc__, file=sys.stderr)
        return 2
    result = check(Path(sys.argv[1]), sys.argv[2])
    print(json.dumps(result))
    return 0


if __name__ == "__main__":
    sys.exit(main())
