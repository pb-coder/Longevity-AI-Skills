from __future__ import annotations

"""Canonical Workout Tracker sheet module.

Single source of truth for:

- **Layout**: headers, column counts, widths (monthly / bodyweight / DB).
- **Coercions**: ``date_str`` for Date cells, ``_numeric_cell`` for
  stringified numbers, ``bw_locate_date`` for the bodyweight row layout.
- **Styling**: fonts, fills, borders, alignment, freeze pane.

Shared by ``/log`` (applied on every append) and ``/maintain`` (full-sheet
restyle). ``/coach`` imports the coercion helpers only. Every
``style_*_sheet`` function is idempotent — running it twice in a row is a
no-op.
"""
from datetime import datetime, date
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------ palette
HEADER_FILL = "FFBDC3C7"
DATA_FILL   = "FFF2F3F4"
NAVY        = "FF2C3E50"
LIGHT_GRAY  = "FFD5D8DC"
SUBSEC_FILL = "FFEAEDED"
BLACK = "FF000000"
WHITE = "FFFFFFFF"
SEP_COLOR = "FFBDC3C7"

font_header    = Font(bold=True, size=10, color=BLACK)
font_section_w = Font(bold=True, size=10, color=WHITE)
font_section_b = Font(bold=True, size=10, color=BLACK)
font_subsec    = Font(bold=True, italic=True, size=10, color=BLACK)
font_data      = Font(bold=False, size=10, color=BLACK)
font_total     = Font(bold=True, size=10, color=BLACK)

fill_header = PatternFill("solid", fgColor=HEADER_FILL)
fill_navy   = PatternFill("solid", fgColor=NAVY)
fill_gray   = PatternFill("solid", fgColor=LIGHT_GRAY)
fill_subsec = PatternFill("solid", fgColor=SUBSEC_FILL)
fill_data   = PatternFill("solid", fgColor=DATA_FILL)
fill_total  = PatternFill("solid", fgColor=LIGHT_GRAY)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
align_left   = Alignment(horizontal="left",   vertical="center", wrap_text=False)

sep_side = Side(style="thin", color=SEP_COLOR)
border_session = Border(top=sep_side)
no_border = Border()

# ------------------------------------------------------------------ structure
# Monthly sheet columns (A..R). A=SESSION (per-month number, merged per date).
# Notes column reads better left-aligned. Volume holds a formula, not a number.
# Laps (col 18) is swim-specific: filled by the Apple importer from
# HKWorkoutEventTypeLap counts and by /log when the user types `<N> laps` /
# `<N> lengths` / `<N> bahnen` on a swim row. None for non-swim rows.
MONTHLY_HEADERS = [
    "SESSION", "Date", "#", "Exercise", "Set", "Reps", "kg", "Volume", "Notes",
    "Distance (km)", "Duration (min)", "Pace (min/km)", "Avg HR",
    "Active Cal", "Total Cal", "Elevation (m)", "Elapsed", "Laps",
]

# Deload marker text — case-sensitive on write (canonical form), but
# substring-matched case-insensitively on read so user-edited variants
# ("deload workout", "DELOAD WORKOUT") still register. Lives on the
# TOTAL row's Notes column for strength sessions.
DELOAD_MARKER_TEXT = "Deload Workout"


def _extract_deload_marker(notes) -> tuple[bool, str | None]:
    """Inspect a Notes cell value and split out the Deload Workout marker.

    Returns ``(deload_present, remaining_notes)``:
    - ``deload_present``: True if ``DELOAD_MARKER_TEXT`` appears (case-insensitive).
    - ``remaining_notes``: the input minus the marker token, with empty
      separator detritus (``"; "``, ``" |"``) tidied. ``None`` if nothing
      meaningful remains.

    Used during the strength-session-metadata-to-TOTAL migration: the
    styler hoists the deload flag to the TOTAL row's Notes while
    preserving any user-written warmup comment on the original row.
    """
    if notes in (None, ""):
        return False, None
    s = str(notes)
    lower = s.lower()
    marker = DELOAD_MARKER_TEXT.lower()
    if marker not in lower:
        return False, s.strip() or None
    # Strip every occurrence + the separator that joined it to user text.
    import re
    pattern = re.compile(re.escape(DELOAD_MARKER_TEXT), re.IGNORECASE)
    cleaned = pattern.sub("", s)
    # Tidy joining separators around the now-removed token.
    cleaned = re.sub(r"\s*;\s*;\s*", "; ", cleaned)
    cleaned = re.sub(r"^\s*[;|,]\s*", "", cleaned)
    cleaned = re.sub(r"\s*[;|,]\s*$", "", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return True, cleaned or None
MONTHLY_WIDTHS = {
    "A": 8,  "B": 12, "C": 5,  "D": 28, "E": 5,  "F": 6, "G": 6,
    "H": 9,  "I": 24, "J": 13, "K": 14, "L": 13, "M": 9,
    "N": 11, "O": 11, "P": 13, "Q": 11, "R": 7,
}
MONTHLY_LEFT_COLS = {"I"}
MONTHLY_COLS = 18
TOTAL_LABEL = "TOTAL"

# Tombstones removed in 2026-05. The importers are now scoped to the
# current calendar month only — see ``_current_month_key`` plus the
# month-gate in ``upsert_monthly_cardio`` / ``upsert_monthly_strength_session``.
# Past months are never re-scanned, so a deleted row stays deleted
# without needing a separate tombstone record.


def _current_month_key(today_d: date | None = None) -> str:
    """Return ``YYYY.MM`` for the current calendar month.

    Single source of truth for the month-gate that bounds where the
    importers can write. ``today_d`` is overridable for tests; production
    callers leave it ``None`` and ``date.today()`` is used.
    """
    d = today_d or date.today()
    return f"{d.year:04d}.{d.month:02d}"


# ------------------------------------------------------------------ helpers
def find_last_data_cell(ws):
    """Return (last_data_row, last_data_col) — the max row/col with any value."""
    last_r, last_c = 0, 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.value != "":
                if cell.row > last_r:
                    last_r = cell.row
                if cell.column > last_c:
                    last_c = cell.column
    return last_r, last_c


# ------------------------------------------------------------------ styling
def _to_num(v):
    """Coerce to float for strength-session detection. Blank/None → 0."""
    if v in (None, ""):
        return 0.0
    try:
        return float(str(v).replace(",", "."))
    except (TypeError, ValueError):
        return 0.0


def date_str(v):
    """Coerce a Date cell value to a canonical ``YYYY-MM-DD`` string.

    Contract:
    - ``None`` or ``""`` → ``None``
    - ``datetime`` / ``date`` → ``"YYYY-MM-DD"``
    - anything else → ``str(v).strip()[:10]`` (covers bare strings, the
      ``"2026-04-20 00:00:00"`` form Excel sometimes emits, and stray
      non-string values).

    Legacy rows imported via Numbers/Excel autoformat can land as
    ``datetime`` objects; mixing types breaks any comparison (sort,
    dict-key merge, equality-based session grouping). Callers coerce
    unconditionally and rely on the ``None`` sentinel for empty cells.
    """
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()[:10]


def _classify_session_rows(rows: list[dict]) -> tuple[list[str], bool]:
    """Classify a session's rows by kind, returning ``(kinds, is_strength)``.

    Per-row classification (preserves the convention used across both the
    monthly-sheet styler and downstream readers):
    - ``strength``: ``kg * reps > 0`` (user lifted weight).
    - ``cardio``:   ``distance > 0`` (Run / Cycle / Hike / Swim — has GPS).
    - ``other``:    no kg*reps and no distance (warmup, HIIT, yoga,
      bodyweight squat, etc.).

    ``is_strength`` is true when any row classifies as ``strength`` — used
    by the styler to decide whether to emit a TOTAL row and whether to
    blank session-level metadata cells on non-cardio rows.
    """
    kinds: list[str] = []
    for rd in rows:
        kg_v = _to_num(rd.get("kg"))
        reps_v = _to_num(rd.get("reps"))
        if kg_v * reps_v > 0:
            kinds.append("strength")
            continue
        dist_v = _to_num(rd.get("distance"))
        if dist_v > 0:
            kinds.append("cardio")
            continue
        kinds.append("other")
    return kinds, ("strength" in kinds)


def _numeric_cell(v):
    """Coerce stringified numbers (incl. European comma decimals like ``"67,5"``)
    to int or float. Returns the original value for anything that isn't purely
    numeric — MM:SS durations, blank cells, text notes — so callers can apply
    this indiscriminately to numeric columns without destroying non-numeric
    entries that happen to share a column historically.

    Excel requires real numbers in cells consumed by arithmetic formulas;
    ``=F*G`` on text returns #VALUE! (see the 2026-04-06 regression where
    legacy data imported kg as string ``"67,5"``). Normalising here keeps the
    sheet arithmetic-safe regardless of how the value was originally logged.
    """
    if v in (None, ""):
        return v
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if not s:
        return v
    try:
        f = float(s.replace(",", "."))
    except ValueError:
        return v
    return int(f) if f.is_integer() else f


# ============================================================ Monthly sheet
def style_monthly_sheet(ws):
    """Apply canonical styling to a YYYY.MM workout log sheet.

    Idempotent. Running twice in a row is a no-op. Responsibilities:

    - Migrate legacy 12-col layout (A="Date") to the 13-col layout by
      inserting a leftmost SESSION column.
    - Normalise Date cells to ``YYYY-MM-DD`` strings (legacy datetime cells
      are coerced), then sort sessions by date ascending and merge any
      same-date sessions that became non-contiguous on disk — e.g. after a
      backfill row got appended at the bottom instead of into its day's
      block.
    - Rebuild SESSION column: per-month session numbers (1..N, chronological
      after the sort), merged vertically across each session's rows.
    - Rebuild TOTAL row at the end of each strength session with
      ``=SUM(H{first}:H{last})`` in the Volume column. Cardio-only sessions
      (every row has kg=0 and reps=0) get no TOTAL row.
    - Rewrite the Volume column on every set row as ``=F{r}*G{r}``
      (reps × kg) — the old hard-coded numeric volumes are discarded.
    - Apply header/data/TOTAL fills, fonts, widths, freeze pane.

    The data-row region is cleared and rebuilt from parsed rows so merges,
    TOTAL row placement, and session numbering can't drift out of sync with
    the data.
    """
    # ---- Migration: legacy layout had A="Date". Shift everything one col right.
    if ws.cell(row=1, column=1).value == "Date":
        ws.insert_cols(1)

    # ---- Unmerge any existing ranges so we can re-apply cleanly.
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))

    # ---- Pass 1: read all data rows into memory, grouped by contiguous date.
    # TOTAL rows are not appended to the session's data rows (they get
    # rebuilt) but their metadata + Notes are captured into the session's
    # ``total_meta`` dict — this is the canonical source for session-level
    # metadata after the move from warmup-row.
    last_r, _ = find_last_data_cell(ws)
    sessions: list[dict] = []
    current: dict | None = None
    for r in range(2, max(last_r, 1) + 1):
        date_val = date_str(ws.cell(row=r, column=2).value)
        ex_val = ws.cell(row=r, column=4).value

        # TOTAL row: harvest session-level metadata for the most recent
        # session before discarding the row. The TOTAL row's date may be
        # blank on legacy data; fall back to ``current["date"]``.
        if ex_val == TOTAL_LABEL:
            if current is not None:
                current["total_meta"] = {
                    "date":         date_val or current.get("date"),
                    "notes":        ws.cell(row=r, column=9).value,
                    "duration":     ws.cell(row=r, column=11).value,
                    "avg_hr":       ws.cell(row=r, column=13).value,
                    "active_cal":   ws.cell(row=r, column=14).value,
                    "total_cal":    ws.cell(row=r, column=15).value,
                    "elevation_m": ws.cell(row=r, column=16).value,
                    "elapsed":      ws.cell(row=r, column=17).value,
                    "laps":         ws.cell(row=r, column=18).value,
                }
            continue
        # Drop fully-empty rows.
        if date_val in (None, "") and ex_val in (None, ""):
            continue

        row_data = {
            "date":      date_val,
            "num":       ws.cell(row=r, column=3).value,
            "exercise":  ex_val,
            "set":       ws.cell(row=r, column=5).value,
            "reps":      ws.cell(row=r, column=6).value,
            "kg":        ws.cell(row=r, column=7).value,
            "notes":     ws.cell(row=r, column=9).value,
            "distance":  ws.cell(row=r, column=10).value,
            "duration":  ws.cell(row=r, column=11).value,
            "pace":      ws.cell(row=r, column=12).value,
            "avg_hr":    ws.cell(row=r, column=13).value,
            "active_cal":  ws.cell(row=r, column=14).value,
            "total_cal":   ws.cell(row=r, column=15).value,
            "elevation_m": ws.cell(row=r, column=16).value,
            "elapsed":     ws.cell(row=r, column=17).value,
            "laps":        ws.cell(row=r, column=18).value,
        }

        if current is None or date_val != current["date"]:
            current = {"date": date_val, "rows": [], "total_meta": {}}
            sessions.append(current)
        current["rows"].append(row_data)

    # ---- Sort sessions by date ascending and merge any that share a date.
    # A shared-date pair happens when a backfill row (earlier than existing
    # data) got appended to the end of the sheet and is no longer contiguous
    # with that date's original block. Without this, backfilled workouts end
    # up stranded at the bottom and SESSION numbers drift out of chronological
    # order.
    merged: dict = {}
    for sess in sessions:
        date = sess["date"]
        slot = merged.setdefault(date, {"date": date, "rows": [], "total_meta": {}})
        slot["rows"].extend(sess["rows"])
        # If two same-date blocks each had a TOTAL row, prefer non-None
        # values from either (same merge semantics as the row-level data).
        for k, v in (sess.get("total_meta") or {}).items():
            if v not in (None, "") and slot["total_meta"].get(k) in (None, ""):
                slot["total_meta"][k] = v
    sessions = sorted(
        merged.values(),
        key=lambda s: (s["date"] is None, s["date"] or ""),
    )

    # ---- Clear data area.
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # ---- Pass 2: write header + data back with SESSION numbers, merges,
    # Volume formulas, TOTAL rows.
    for c, label in enumerate(MONTHLY_HEADERS, 1):
        ws.cell(row=1, column=c, value=label)

    write_row = 2
    for session_num, sess in enumerate(sessions, start=1):
        first_row = write_row

        # Per-row + per-session classification via the shared helper.
        # Cols 11 (Duration), 13 (Avg HR), 14-17 placement:
        # - **Strength session**: session metadata lives on the TOTAL row
        #   only. All non-cardio rows (warmup + working sets) have those
        #   cells blank. Cardio rows mixed into a strength day keep their
        #   own per-row metadata (each Apple-recorded ride is independent).
        # - **Pure cardio session** (no strength rows): no TOTAL row;
        #   every row keeps its own per-row metadata.
        kinds, is_strength = _classify_session_rows(sess["rows"])

        # Hoist candidates for TOTAL row metadata. Priority order:
        #   1. The TOTAL row's existing values (canonical post-migration).
        #   2. Any non-cardio data row's value (legacy migration path,
        #      taken from the warmup row before this change).
        total_meta_in = sess.get("total_meta") or {}
        hoist_duration = total_meta_in.get("duration")
        hoist_avg_hr = total_meta_in.get("avg_hr")
        hoist_active = total_meta_in.get("active_cal")
        hoist_total = total_meta_in.get("total_cal")
        hoist_elev = total_meta_in.get("elevation_m")
        hoist_elapsed = total_meta_in.get("elapsed")
        # Hoisted Notes from TOTAL: split off the Deload Workout marker
        # so it survives independently of any user-added text on TOTAL.
        total_notes_in = total_meta_in.get("notes")
        deload_present, total_notes_remainder = _extract_deload_marker(total_notes_in)

        if is_strength:
            for i, rd in enumerate(sess["rows"]):
                if kinds[i] == "cardio":
                    continue
                if hoist_duration in (None, "") and rd.get("duration") not in (None, ""):
                    hoist_duration = rd["duration"]
                if hoist_avg_hr in (None, "") and rd.get("avg_hr") not in (None, ""):
                    hoist_avg_hr = _numeric_cell(rd["avg_hr"])
                if hoist_active in (None, "") and rd.get("active_cal") not in (None, ""):
                    hoist_active = _numeric_cell(rd["active_cal"])
                if hoist_total in (None, "") and rd.get("total_cal") not in (None, ""):
                    hoist_total = _numeric_cell(rd["total_cal"])
                if hoist_elev in (None, "") and rd.get("elevation_m") not in (None, ""):
                    hoist_elev = _numeric_cell(rd["elevation_m"])
                if hoist_elapsed in (None, "") and rd.get("elapsed") not in (None, ""):
                    hoist_elapsed = rd["elapsed"]
                # Inspect this row's Notes for any deload marker —
                # legacy /log convention put it on the warmup row. Strip
                # it out so only the user's per-exercise text remains;
                # raise the deload flag for the TOTAL row.
                row_deload, row_notes_remainder = _extract_deload_marker(rd.get("notes"))
                if row_deload:
                    deload_present = True
                    rd["notes"] = row_notes_remainder

        for idx, rd in enumerate(sess["rows"]):
            ws.cell(row=write_row, column=1, value=session_num)
            ws.cell(row=write_row, column=2, value=rd["date"])
            ws.cell(row=write_row, column=3, value=_numeric_cell(rd["num"]))
            ws.cell(row=write_row, column=4, value=rd["exercise"])
            ws.cell(row=write_row, column=5, value=_numeric_cell(rd["set"]))
            ws.cell(row=write_row, column=6, value=_numeric_cell(rd["reps"]))
            ws.cell(row=write_row, column=7, value=_numeric_cell(rd["kg"]))
            ws.cell(row=write_row, column=8, value=f"=F{write_row}*G{write_row}")
            ws.cell(row=write_row, column=9, value=rd["notes"])
            ws.cell(row=write_row, column=10, value=_numeric_cell(rd["distance"]))
            ws.cell(row=write_row, column=12, value=rd["pace"])
            if is_strength and kinds[idx] == "cardio":
                # Cardio row inside a mixed-session day — keep its own
                # per-row Duration / Avg HR / cols 14-18.
                ws.cell(row=write_row, column=11, value=rd["duration"])
                ws.cell(row=write_row, column=13, value=_numeric_cell(rd["avg_hr"]))
                ws.cell(row=write_row, column=14,
                        value=_numeric_cell(rd.get("active_cal")))
                ws.cell(row=write_row, column=15,
                        value=_numeric_cell(rd.get("total_cal")))
                ws.cell(row=write_row, column=16,
                        value=_numeric_cell(rd.get("elevation_m")))
                ws.cell(row=write_row, column=17, value=rd.get("elapsed"))
                ws.cell(row=write_row, column=18, value=_numeric_cell(rd.get("laps")))
            elif is_strength:
                # Strength/other row — session metadata moved to TOTAL,
                # cells stay blank.
                ws.cell(row=write_row, column=11).value = None
                ws.cell(row=write_row, column=13).value = None
                ws.cell(row=write_row, column=14).value = None
                ws.cell(row=write_row, column=15).value = None
                ws.cell(row=write_row, column=16).value = None
                ws.cell(row=write_row, column=17).value = None
                ws.cell(row=write_row, column=18).value = None
            else:
                # Pure cardio session — each row keeps its own metadata.
                ws.cell(row=write_row, column=11, value=rd["duration"])
                ws.cell(row=write_row, column=13, value=_numeric_cell(rd["avg_hr"]))
                ws.cell(row=write_row, column=14,
                        value=_numeric_cell(rd.get("active_cal")))
                ws.cell(row=write_row, column=15,
                        value=_numeric_cell(rd.get("total_cal")))
                ws.cell(row=write_row, column=16,
                        value=_numeric_cell(rd.get("elevation_m")))
                ws.cell(row=write_row, column=17, value=rd.get("elapsed"))
                ws.cell(row=write_row, column=18, value=_numeric_cell(rd.get("laps")))
            write_row += 1
        last_set_row = write_row - 1

        if is_strength:
            # TOTAL row: Date, Volume formula, all session-level metadata,
            # and the Deload Workout marker on Notes when present. Merge
            # the SESSION column (col 1) through this row alongside the
            # data rows below. Laps is row-level (swim only) and never
            # rolls up to TOTAL — strength sessions don't have laps.
            ws.cell(row=write_row, column=1, value=session_num)
            ws.cell(row=write_row, column=2, value=sess["date"])
            ws.cell(row=write_row, column=4, value=TOTAL_LABEL)
            ws.cell(row=write_row, column=8, value=f"=SUM(H{first_row}:H{last_set_row})")
            ws.cell(row=write_row, column=9,
                    value=DELOAD_MARKER_TEXT if deload_present else None)
            ws.cell(row=write_row, column=11, value=hoist_duration)
            ws.cell(row=write_row, column=13, value=hoist_avg_hr)
            ws.cell(row=write_row, column=14, value=hoist_active)
            ws.cell(row=write_row, column=15, value=hoist_total)
            ws.cell(row=write_row, column=16, value=hoist_elev)
            ws.cell(row=write_row, column=17, value=hoist_elapsed)
            ws.cell(row=write_row, column=18).value = None
            merge_last = write_row
            write_row += 1
        else:
            merge_last = last_set_row

        if merge_last > first_row:
            ws.merge_cells(
                start_row=first_row, end_row=merge_last,
                start_column=1, end_column=1,
            )

    last_written = write_row - 1

    # ---- Styling: header.
    for c in range(1, MONTHLY_COLS + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = no_border

    # ---- Styling: data rows + TOTAL rows. Skip merged follower cells —
    # openpyxl MergedCell objects don't accept style writes.
    for r in range(2, last_written + 1):
        is_total = ws.cell(row=r, column=4).value == TOTAL_LABEL
        for c in range(1, MONTHLY_COLS + 1):
            cell = ws.cell(row=r, column=c)
            if cell.__class__.__name__ == "MergedCell":
                continue
            cell.font = font_total if is_total else font_data
            cell.fill = fill_total if is_total else fill_data
            col = get_column_letter(c)
            cell.alignment = align_left if col in MONTHLY_LEFT_COLS else align_center
            cell.border = no_border

    # ---- Column widths + freeze pane.
    for col, w in MONTHLY_WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"



# ================================================== Auto-cardio: monthly append
_MONTH_RE_STR = r"^\d{4}\.\d{2}$"


def _is_monthly_sheet(name: str) -> bool:
    import re as _re
    return bool(_re.match(_MONTH_RE_STR, name))


def canonicalize_sheet_order(wb) -> None:
    """Re-order monthly sheets newest → oldest. Idempotent.

    Post-PR1 the workbook holds only ``YYYY.MM`` monthly sheets — Health
    Metrics / Workout Sessions / Profile / Bodyweight / Exercises
    Database all moved to ``<person>/data/`` CSVs. Any unexpected
    non-monthly sheet sticks at the end so ``/maintain`` can flag it.
    """
    names = list(wb.sheetnames)
    months = sorted((n for n in names if _is_monthly_sheet(n)), reverse=True)
    other = [n for n in names if not _is_monthly_sheet(n)]
    desired = months + other

    for target_idx, name in enumerate(desired):
        sheet_obj = wb[name]
        cur_idx = wb.sheetnames.index(name)
        if cur_idx != target_idx:
            wb.move_sheet(sheet_obj, offset=target_idx - cur_idx)


def _format_duration_mmss(duration_min) -> str | None:
    """Coerce a numeric or MM:SS duration to a canonical ``MM:SS`` string.

    - ``None`` / ``""`` → None
    - already a ``MM:SS`` string → returned verbatim (after a defensive
      sanity check that minutes are non-negative)
    - any numeric value (or numeric-looking string, incl. European comma) →
      formatted as ``MM:SS``
    Auto-cardio rows previously wrote raw float minutes (e.g. ``60.6``);
    this helper unifies them with manual-log MM:SS strings.
    """
    if duration_min in (None, ""):
        return None
    if isinstance(duration_min, str):
        s = duration_min.strip()
        if ":" in s:
            return s if s else None
        try:
            duration_min = float(s.replace(",", "."))
        except ValueError:
            return None
    try:
        f = float(duration_min)
    except (TypeError, ValueError):
        return None
    if f <= 0:
        return None
    whole = int(f)
    secs = int(round((f - whole) * 60))
    if secs == 60:
        whole += 1
        secs = 0
    return f"{whole}:{secs:02d}"


# Pace sanity bounds. ``< 0.5 min/km`` is faster than any sustained human
# (Bolt's 100 m WR ≈ 2:35/km); ``> 60 min/km`` is slower than a slow
# stroll. Either side of this range is almost always a unit-bug or a
# distance/duration entry mismatch — bail out and surface as blank so the
# bad row stands out instead of a silent ``0:01``.
PACE_MIN_PER_KM_LOWER = 0.5
PACE_MIN_PER_KM_UPPER = 60.0


def _format_pace_min_per_km(duration_min: float | None,
                            distance_km: float | None) -> str | None:
    """Return a ``MM:SS`` per-km pace string, or None if not computable.

    Mirrors the manual-log pace format. We only emit a pace when both
    distance and duration are positive — bare-duration cardio (HIIT,
    swim laps without distance) leaves the cell blank. Pace values
    outside ``[0.5, 60]`` min/km also return None — see the bounds
    constants above for rationale.
    """
    if not duration_min or not distance_km:
        return None
    if duration_min <= 0 or distance_km <= 0:
        return None
    pace_min = duration_min / distance_km
    if pace_min < PACE_MIN_PER_KM_LOWER or pace_min > PACE_MIN_PER_KM_UPPER:
        return None
    whole = int(pace_min)
    secs = int(round((pace_min - whole) * 60))
    if secs == 60:
        whole += 1
        secs = 0
    return f"{whole}:{secs:02d}"


# Tolerance used when matching an Apple workout to an existing manual cardio
# row on the same date + exercise. Apple sometimes records 28.0 min for what
# the user logged as 30 min; ±1 min absorbs the rounding without merging
# genuinely different sessions (a manual 25-min run and an Apple 35-min run
# on the same day are different workouts).
CARDIO_DUPLICATE_DURATION_TOLERANCE_MIN = 1.0

AUTO_IMPORT_NOTE = "auto-imported from Apple"

# When elapsed and active duration differ by less than this many minutes the
# elapsed-time field is dropped from the auto-cardio note (no meaningful pause
# happened, so showing both adds noise). Threshold matches the human eye —
# 1 min of stop time on a 30-min run isn't worth surfacing.
ELAPSED_TIME_MIN_DELTA = 1.0


def _format_elapsed_hms(elapsed_min: float | None) -> str | None:
    """Render elapsed minutes as ``H:MM:SS`` for long activities, else ``MM:SS``.

    Used by the auto-cardio note builder. Returns None on missing/zero input
    so the caller can omit the field entirely.
    """
    if elapsed_min in (None, "") or not isinstance(elapsed_min, (int, float)):
        return None
    if elapsed_min <= 0:
        return None
    total_seconds = int(round(elapsed_min * 60))
    hours, rem = divmod(total_seconds, 3600)
    minutes, seconds = divmod(rem, 60)
    if hours > 0:
        return f"{hours}:{minutes:02d}:{seconds:02d}"
    return f"{minutes}:{seconds:02d}"


def build_auto_cardio_note(
    *,
    active_cal: float | None = None,
    total_cal: float | None = None,
    elevation_m: float | None = None,
    duration_min: float | None = None,
    elapsed_min: float | None = None,
) -> str:
    """Build the structured Notes string for an auto-imported cardio row.

    Format: ``auto-imported from Apple | active 753 kcal | total 1041 kcal |
    elevation 73 m | elapsed 3:53:23``

    Each ``| field value`` is dropped when the source value is None or
    zero. ``elapsed`` is dropped when it differs from ``duration_min`` by
    less than ``ELAPSED_TIME_MIN_DELTA`` minutes (no meaningful pauses).
    Single source of truth — both importers and the backfill script call
    this helper so the format stays consistent.
    """
    parts: list[str] = [AUTO_IMPORT_NOTE]

    if active_cal not in (None, 0, 0.0):
        parts.append(f"active {int(round(active_cal))} kcal")
    if total_cal not in (None, 0, 0.0):
        # Skip the total field if it's not meaningfully larger than active —
        # otherwise we'd emit identical-looking active/total values.
        if active_cal is None or total_cal - active_cal >= 1:
            parts.append(f"total {int(round(total_cal))} kcal")
    if elevation_m not in (None, 0, 0.0):
        parts.append(f"elevation {int(round(elevation_m))} m")

    if elapsed_min not in (None, 0, 0.0):
        delta = abs(elapsed_min - (duration_min or 0))
        if duration_min is None or delta >= ELAPSED_TIME_MIN_DELTA:
            elapsed_str = _format_elapsed_hms(elapsed_min)
            if elapsed_str:
                parts.append(f"elapsed {elapsed_str}")

    return " | ".join(parts)


def _parse_duration_minutes(v):
    """Coerce a Duration-column value to float minutes.

    Manual logs use MM:SS strings (``"60:33"`` for one hour and 33s).
    Auto-imports use floats (``60.6``). Any other shape returns None so the
    caller's dedupe falls back to a date+exercise match (manual-wins).
    """
    if v in (None, ""):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    if ":" in s:
        parts = s.split(":")
        try:
            mm = int(parts[0])
            ss = int(parts[1]) if len(parts) > 1 else 0
            return mm + ss / 60.0
        except ValueError:
            return None
    try:
        return float(s.replace(",", "."))
    except ValueError:
        return None


def upsert_monthly_cardio(wb, rows: list[dict], allow_past_months: bool = False) -> list[str]:
    """Append cardio rows to ``YYYY.MM`` sheets with dedupe + restyle.

    ``rows`` is a list of dicts with keys:

      ``date`` (YYYY-MM-DD), ``exercise`` (canonical name from
      ``apple_workout_types.APPLE_TO_TRACKER_EXERCISE``),
      ``duration_min``, ``distance_km``, ``avg_hr``.

    For each row:

    - The target sheet is the ``YYYY.MM`` matching the row's date. Created
      with the canonical 13-col header if missing.
    - Dedupe rule: skip if the sheet already has a row matching ALL of:
      same date, same exercise (case-insensitive), and duration within
      ±1.0 min. Manual entries (no ``auto-imported from Apple`` note) win
      — never overwritten. A previously auto-imported row with the same
      key is also a no-op (idempotency).
    - Otherwise the new row is appended at the bottom; the styler
      (called once per touched sheet at the end) sorts and rewrites it
      into chronological order, populates ``SESSION``, and writes the
      ``=F*G`` Volume formula. The note column carries
      ``"auto-imported from Apple"`` so the manual-wins rule can find it.

    ``allow_past_months`` (default False) bypasses the current-month gate
    and lets rows flow into prior YYYY.MM sheets too. Use sparingly — the
    gate is the whole reason past months are immune to re-scan drift.
    Intended for one-off backfills (e.g. enabling auto-cardio retroactively
    for a tracker that had it off).

    Returns one summary string per touched sheet, plus a roll-up. Empty
    input returns a no-op summary.
    """
    if not rows:
        return ["Auto-cardio: 0 rows considered"]

    # Bucket incoming rows by ``YYYY.MM`` once so we touch each sheet exactly
    # once (read existing → append new → restyle), regardless of how rows
    # were ordered.
    #
    # Current-month gate: importers only ever write into the current
    # calendar month's monthly sheet. Past months are "finished" and are
    # never re-scanned, so a row the user deleted from 2026.02 stays
    # deleted on the next import without needing a tombstone. This is the
    # whole reason the Tombstones sheet was removed in 2026-05.
    # ``allow_past_months=True`` opts out for one-off backfills.
    current_month = _current_month_key()
    skipped_past_month = 0
    by_month: dict[str, list[dict]] = {}
    for r in rows:
        d = str(r.get("date") or "")[:10]
        if not d or len(d) != 10:
            continue
        key = f"{d[:4]}.{d[5:7]}"
        if not allow_past_months and key != current_month:
            skipped_past_month += 1
            continue
        by_month.setdefault(key, []).append(r)

    if not by_month:
        if skipped_past_month:
            return [
                f"Auto-cardio: 0 rows considered "
                f"({skipped_past_month} skipped — past months are not re-scanned)"
            ]
        return ["Auto-cardio: 0 rows considered (no valid dates)"]

    summaries: list[str] = []
    total_appended = 0
    total_skipped = 0

    for month_key in sorted(by_month.keys()):
        month_rows = by_month[month_key]
        if month_key in wb.sheetnames:
            ws = wb[month_key]
            sheet_created = False
        else:
            ws = wb.create_sheet(title=month_key)
            for c, label in enumerate(MONTHLY_HEADERS, 1):
                ws.cell(row=1, column=c, value=label)
            sheet_created = True

        # Snapshot existing rows for dedupe lookup. Index by
        # (date, lowercased exercise) → list of (row_index, duration_min,
        # is_auto) so the manual-wins rule and the sparse-merge metadata
        # update both have access to the source row.
        existing_index: dict[tuple, list[tuple]] = {}
        last_r, _ = find_last_data_cell(ws)
        for r in range(2, last_r + 1):
            date_v = date_str(ws.cell(row=r, column=2).value)
            ex_v = ws.cell(row=r, column=4).value
            if not date_v or not ex_v:
                continue
            ex_str = str(ex_v).strip()
            if not ex_str or ex_str.upper() == TOTAL_LABEL:
                continue
            dur_f = _parse_duration_minutes(ws.cell(row=r, column=11).value)
            notes_v = ws.cell(row=r, column=9).value
            # Case-insensitive substring check — both sides lowered.
            is_auto = AUTO_IMPORT_NOTE.lower() in str(notes_v or "").lower()
            key = (date_v, ex_str.lower())
            existing_index.setdefault(key, []).append((r, dur_f, is_auto))

        appended = 0
        skipped_dup = 0
        refreshed = 0
        # Track which existing rows have already been claimed by an input
        # workout so two near-duration Apple workouts can't both match the
        # same tracker row. Without this, fuzzy ±1-min matching causes
        # oscillation when Apple has more workouts than the tracker rows
        # for a date (e.g. 5 cycling segments vs 4 rows).
        claimed_rows: set = set()

        # Determine where to start writing — append after the last data row
        # the styler will sort+merge afterwards anyway. But if the sheet
        # contains TOTAL rows or merged SESSION cells, ``ws.max_row`` is the
        # safer bottom anchor.
        write_row = max(ws.max_row + 1, 2)

        for r in month_rows:
            d = str(r.get("date") or "")[:10]
            ex = r.get("exercise")
            if not d or not ex:
                continue
            ex_lower = ex.strip().lower()
            dur = r.get("duration_min")
            try:
                dur_f = float(dur) if dur is not None else None
            except (TypeError, ValueError):
                dur_f = None

            # Dedupe rule:
            #   - Any existing MANUAL row on the same date + exercise wins
            #     unconditionally — Apple frequently splits a single workout
            #     into multiple short segments (a 0.6-min preamble + the
            #     main 159-min session, both tagged Hiking), and the user
            #     only logged the workout once. Manual-wins is the right
            #     behavior here even though it discards the auxiliary
            #     segments.
            #   - Otherwise: an existing AUTO row with the same date +
            #     exercise + duration (±1 min tolerance) is the
            #     idempotency path. Re-imports of the same export must
            #     produce zero appends.
            matches = existing_index.get((d, ex_lower), [])
            has_manual_match = any(
                (not is_auto) and (er not in claimed_rows)
                for er, _dur, is_auto in matches
            )
            matched_auto_row = None
            if not has_manual_match:
                # Pick the unclaimed auto row with the smallest duration
                # difference within tolerance — stable across re-runs and
                # prevents two near-duration Apple workouts from claiming
                # the same row.
                best = None
                best_diff = None
                for er, existing_dur, is_auto in matches:
                    if not is_auto or er in claimed_rows:
                        continue
                    if existing_dur is None or dur_f is None:
                        diff = 0.0
                    else:
                        diff = abs(existing_dur - dur_f)
                        if diff > CARDIO_DUPLICATE_DURATION_TOLERANCE_MIN:
                            continue
                    if best_diff is None or diff < best_diff:
                        best, best_diff = er, diff
                matched_auto_row = best
            if matched_auto_row is not None:
                claimed_rows.add(matched_auto_row)
                # Existing auto row — refresh stale cols 14-17 in place.
                # Historical importer bug copied the same Notes string to
                # every cardio row of a date; this sparse-merge corrects
                # those values from per-workout Apple data. Only updates
                # cells that are empty OR diverge from incoming by >=5%
                # (the manual-wins / drift threshold from Q6).
                incoming_meta = {
                    14: int(round(r.get("active_cal"))) if isinstance(r.get("active_cal"), (int, float)) and r.get("active_cal") else None,
                    15: int(round(r.get("total_cal"))) if isinstance(r.get("total_cal"), (int, float)) and r.get("total_cal") else None,
                    16: int(round(r.get("elevation_m"))) if isinstance(r.get("elevation_m"), (int, float)) and r.get("elevation_m") else None,
                    17: _format_elapsed_hms(r.get("elapsed_min")),
                    18: int(r.get("laps")) if isinstance(r.get("laps"), (int, float)) and r.get("laps") else None,
                }
                row_changed = False
                for col, new_val in incoming_meta.items():
                    if new_val in (None, ""):
                        continue
                    existing = ws.cell(row=matched_auto_row, column=col).value
                    if existing in (None, ""):
                        ws.cell(row=matched_auto_row, column=col).value = new_val
                        row_changed = True
                    elif _strength_metadata_drifts(existing, new_val):
                        ws.cell(row=matched_auto_row, column=col).value = new_val
                        row_changed = True
                if row_changed:
                    refreshed += 1
                else:
                    skipped_dup += 1
                continue
            if has_manual_match:
                skipped_dup += 1
                continue

            distance = r.get("distance_km")
            avg_hr = r.get("avg_hr")
            pace = _format_pace_min_per_km(dur_f, distance)

            # Compute the next free per-date # for this exercise. Manual
            # rows usually have a #; cardio is typically solo-per-day so
            # this lands at 1, but reuses an existing day's exercise count
            # if the user logged strength earlier the same date.
            existing_nums = []
            for er in range(2, ws.max_row + 1):
                if date_str(ws.cell(row=er, column=2).value) == d:
                    n = ws.cell(row=er, column=3).value
                    if isinstance(n, (int, float)):
                        existing_nums.append(int(n))
            next_num = (max(existing_nums) + 1) if existing_nums else 1

            ws.cell(row=write_row, column=1, value=None)        # SESSION (styler fills)
            ws.cell(row=write_row, column=2, value=d)
            ws.cell(row=write_row, column=3, value=next_num)
            ws.cell(row=write_row, column=4, value=ex)
            ws.cell(row=write_row, column=5, value=1)           # Set
            ws.cell(row=write_row, column=6, value=None)        # Reps (cardio: blank)
            ws.cell(row=write_row, column=7, value=None)        # kg
            ws.cell(row=write_row, column=8, value=f"=F{write_row}*G{write_row}")
            machine_tag = r.get("machine_tag")
            note_value = (
                f"{AUTO_IMPORT_NOTE} | source: {machine_tag}"
                if machine_tag else AUTO_IMPORT_NOTE
            )
            ws.cell(row=write_row, column=9, value=note_value)
            ws.cell(row=write_row, column=10, value=_numeric_cell(distance))
            ws.cell(row=write_row, column=11, value=_format_duration_mmss(dur_f))
            ws.cell(row=write_row, column=12, value=pace)
            ws.cell(row=write_row, column=13, value=_numeric_cell(avg_hr))
            # Apple-watch session metadata in cols 14-17 (single-row cardio
            # session — same row carries the metadata and the workout data).
            ac = r.get("active_cal")
            tc = r.get("total_cal")
            el = r.get("elevation_m")
            ws.cell(row=write_row, column=14,
                    value=int(round(ac)) if isinstance(ac, (int, float)) and ac else None)
            ws.cell(row=write_row, column=15,
                    value=int(round(tc)) if isinstance(tc, (int, float)) and tc else None)
            ws.cell(row=write_row, column=16,
                    value=int(round(el)) if isinstance(el, (int, float)) and el else None)
            ws.cell(row=write_row, column=17,
                    value=_format_elapsed_hms(r.get("elapsed_min")))
            ws.cell(row=write_row, column=18, value=_numeric_cell(r.get("laps")))

            # Track the new row in the dedupe index too so subsequent input
            # rows don't double-add for the same date.
            existing_index.setdefault((d, ex_lower), []).append((write_row, dur_f, True))

            write_row += 1
            appended += 1

        if appended or refreshed:
            style_monthly_sheet(ws)

        total_appended += appended
        total_skipped += skipped_dup
        tag = " (new sheet)" if sheet_created else ""
        refreshed_tag = f", {refreshed} refreshed" if refreshed else ""
        summaries.append(
            f"{month_key}{tag}: {appended} cardio rows appended, "
            f"{skipped_dup} skipped (already present){refreshed_tag}"
        )

    # Re-canonicalize tab order so newly-created month sheets land in the
    # correct position (newest first, after Workout Sessions). Without this,
    # a fresh month sits at the right end of the strip until /maintain runs.
    canonicalize_sheet_order(wb)

    summaries.append(
        f"Auto-cardio total: {total_appended} appended, "
        f"{total_skipped} skipped across {len(by_month)} month(s)"
    )
    if skipped_past_month:
        summaries.append(
            f"Auto-cardio: {skipped_past_month} input rows skipped "
            f"(dated outside the current month {current_month})"
        )
    return summaries


# Threshold for the manual-wins / sparse-update rule on strength session
# metadata cells. If an existing cell value differs from the incoming Apple
# value by < 5%, treat them as the same (idempotency / Apple jitter). If
# the difference is >= 5%, the existing value is treated as user-edited and
# the importer skips with a warning rather than overwriting.
STRENGTH_METADATA_DRIFT_THRESHOLD = 0.05


# ================================================== Strength upsert (monthly)
def _strength_metadata_drifts(existing, incoming) -> bool:
    """Return True if the two values disagree by >= 5% (manual-wins guard).

    Both arguments are numeric (int / float) or Elapsed strings. Strings
    are compared by re-parsing back to minutes via ``_parse_duration_minutes``
    so ``"1:29:18"`` and ``"1:29:20"`` don't trip on the second-level noise.
    None/blank existing → never drifts (fill it in). Equal values → no drift.
    """
    if existing in (None, ""):
        return False
    if incoming in (None, ""):
        return False
    if isinstance(existing, str) or isinstance(incoming, str):
        e_min = _parse_duration_minutes(existing)
        i_min = _parse_duration_minutes(incoming)
        if e_min is None or i_min is None:
            return str(existing).strip() != str(incoming).strip()
        existing_f, incoming_f = e_min, i_min
    else:
        try:
            existing_f = float(existing)
            incoming_f = float(incoming)
        except (TypeError, ValueError):
            return existing != incoming
    if existing_f == 0 and incoming_f == 0:
        return False
    denom = max(abs(existing_f), abs(incoming_f), 1e-9)
    return abs(existing_f - incoming_f) / denom >= STRENGTH_METADATA_DRIFT_THRESHOLD


def upsert_monthly_strength_session(wb, sessions: list[dict]) -> list[str]:
    """Write Apple-watch session metadata onto the first row of each
    matching strength session in the monthly sheets.

    ``sessions`` is a list of dicts:

      ``date`` (YYYY-MM-DD), ``active_cal``, ``total_cal``, ``elevation_m``
      (usually None for indoor strength), ``elapsed_min`` (numeric minutes;
      formatted to ``H:MM:SS`` or ``MM:SS`` on write), ``avg_hr`` (per-cluster
      duration-weighted average heart rate; XML only — HL doesn't carry
      per-workout HR).

    Behavior per session:
    - Locate the YYYY.MM sheet for the date. If missing, skip (the user
      hasn't logged this strength session yet — Apple ahead of manual).
    - Find the first non-TOTAL data row whose Date matches.
    - For each of the 4 metadata cells, apply the manual-wins / drift rule:
      * empty cell → fill with incoming value
      * existing value within 5% of incoming → no-op (idempotency)
      * existing value diverges >= 5% → skip that cell with a warning
    - After touching any cell, re-style the sheet via ``style_monthly_sheet``
      so the convention is enforced uniformly.

    Returns a list of human-readable summary lines, including any
    drift warnings, so the importer can surface them.
    """
    if not sessions:
        return ["Strength sessions: 0 considered"]

    summaries: list[str] = []
    written = 0
    skipped_no_match = 0
    skipped_no_change = 0
    skipped_past_month = 0
    drift_warnings: list[str] = []

    touched_sheets: set = set()

    # Current-month gate: the importer never writes session metadata
    # into past-month sheets. A session logged in 2026.04 stays at
    # whatever metadata the user logged manually; April is "finished".
    current_month = _current_month_key()

    for sess in sessions:
        d = str(sess.get("date") or "")[:10]
        if not d or len(d) != 10:
            continue
        month_key = f"{d[:4]}.{d[5:7]}"
        if month_key != current_month:
            skipped_past_month += 1
            continue
        if month_key not in wb.sheetnames:
            skipped_no_match += 1
            continue
        ws = wb[month_key]

        # Locate the TOTAL row for this date. The styler emits a TOTAL
        # row for every strength session (any kg*reps>0 row); session
        # metadata lives there. The TOTAL row's Date (col 2) was added
        # in this migration; legacy rows may still have a blank Date,
        # so fall back to the row directly above when the date matches.
        target_row = None
        last_r, _ = find_last_data_cell(ws)
        date_seen_above = False
        for r in range(2, last_r + 1):
            row_date = date_str(ws.cell(row=r, column=2).value)
            ex_val = ws.cell(row=r, column=4).value
            ex_str = str(ex_val).strip() if ex_val else ""
            if ex_str.upper() == TOTAL_LABEL:
                if row_date == d or (row_date in (None, "") and date_seen_above):
                    target_row = r
                    break
                date_seen_above = False
                continue
            if row_date == d:
                date_seen_above = True
            elif row_date not in (None, ""):
                date_seen_above = False

        if target_row is None:
            # No TOTAL row for this date — either the user hasn't logged
            # the strength session yet (Apple ahead of /log), or the
            # session has no kg*reps>0 row (the styler skipped TOTAL).
            skipped_no_match += 1
            continue

        ac = sess.get("active_cal")
        tc = sess.get("total_cal")
        el = sess.get("elevation_m")
        em = sess.get("elapsed_min")
        ah = sess.get("avg_hr")
        dur = sess.get("duration_min")

        incoming = {
            11: _format_duration_mmss(dur),  # Duration (col K) — active workout time, MM:SS
            13: round(float(ah), 1) if isinstance(ah, (int, float)) and ah else None,
            14: int(round(ac)) if isinstance(ac, (int, float)) and ac else None,
            15: int(round(tc)) if isinstance(tc, (int, float)) and tc else None,
            16: int(round(el)) if isinstance(el, (int, float)) and el else None,
            17: _format_elapsed_hms(em),
        }

        any_change = False
        for col, new_val in incoming.items():
            if new_val in (None, ""):
                continue
            existing = ws.cell(row=target_row, column=col).value
            if existing in (None, ""):
                ws.cell(row=target_row, column=col, value=new_val)
                any_change = True
            elif _strength_metadata_drifts(existing, new_val):
                drift_warnings.append(
                    f"  - {d} col {col}: kept manual value {existing!r} "
                    f"(Apple reports {new_val!r}, differs >=5%)"
                )
            # else: within tolerance → no-op (idempotency)

        if any_change:
            written += 1
            touched_sheets.add(month_key)
        else:
            skipped_no_change += 1

    for sheet_name in sorted(touched_sheets):
        style_monthly_sheet(wb[sheet_name])

    summaries.append(
        f"Strength sessions: {written} written, "
        f"{skipped_no_change} no-op (already up to date), "
        f"{skipped_no_match} skipped (no matching session row)"
    )
    if skipped_past_month:
        summaries.append(
            f"Strength sessions: {skipped_past_month} dated outside the "
            f"current month {current_month} — past months are not re-scanned"
        )
    if drift_warnings:
        summaries.append(
            f"Strength sessions: {len(drift_warnings)} manual-wins warnings:"
        )
        summaries.extend(drift_warnings)

    return summaries


