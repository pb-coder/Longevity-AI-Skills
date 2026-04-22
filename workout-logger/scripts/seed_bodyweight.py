"""One-shot historical import for the Bodyweight sheet.

Takes a JSON array of entries and writes them into the Bodyweight sheet
(creating the sheet if missing), deduplicates by date (last-write-wins),
sorts DESCENDING (newest at the top), and applies canonical styling
(Date | Kg | Notes).

Input JSON schema:
    [
      {"date": "2026-01-14", "kg": 75.5, "notes": ""},
      {"date": "2023-12-20", "kg": 64.0, "notes": "evening, not fasted"},
      ...
    ]

Usage:
    python3 seed_bodyweight.py <tracker_path> <entries_json_path>
    python3 seed_bodyweight.py <tracker_path> -   # read JSON from stdin

This is an import tool, not part of the normal /log flow. For ongoing
daily captures, append_workout.py handles writes via the `bodyweight`
payload key.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "shared"))
from sheet_styles import BODYWEIGHT_HEADERS, style_bodyweight_sheet  # noqa: E402


def normalize(entry: dict) -> dict:
    date = str(entry["date"])[:10]
    kg = float(entry["kg"])
    notes = entry.get("notes") or None
    return {"date": date, "kg": kg, "notes": notes}


def ensure_bodyweight_sheet(wb):
    if "Bodyweight" in wb.sheetnames:
        return wb["Bodyweight"]
    ws = wb.create_sheet(title="Bodyweight")
    for col, header in enumerate(BODYWEIGHT_HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)
    return ws


def _date_from_row(row: tuple) -> str | None:
    """Find a YYYY-MM-DD value in the first two columns, regardless of layout.

    Handles both the current 3-col layout (Date | Kg | Notes) and the
    legacy 4-col layout (Year | Date | Kg | Notes) so migrations stay safe.
    """
    for v in row[:2]:
        if v is None or v == "":
            continue
        s = str(v)[:10]
        if len(s) == 10 and s[4] == "-" and s[7] == "-":
            return s
    return None


def read_existing(ws) -> dict[str, dict]:
    """Return {date: {kg, notes}} for existing rows so new entries can merge."""
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        date = _date_from_row(row)
        if date is None:
            continue
        # Kg is the column right after Date. In the current 3-col layout
        # that's index 1; in the legacy 4-col layout it's index 2. Detect
        # by locating the date first.
        date_idx = next(
            (i for i, v in enumerate(row[:2]) if v is not None and str(v)[:10] == date),
            None,
        )
        if date_idx is None:
            continue
        kg_raw = row[date_idx + 1] if len(row) > date_idx + 1 else None
        notes = row[date_idx + 2] if len(row) > date_idx + 2 else None
        try:
            kg = float(kg_raw) if kg_raw is not None else None
        except (TypeError, ValueError):
            continue
        if kg is None:
            continue
        out[date] = {"kg": kg, "notes": notes}
    return out


def write_sorted(ws, merged: dict[str, dict]):
    """Clear old data and rewrite in DESCENDING-date order.

    Unmerges any existing ranges first so the row-delete path can't collide
    with merged cells left over from a prior styling pass (including legacy
    per-year merges on column A).
    """
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    for i, date in enumerate(sorted(merged.keys(), reverse=True), start=2):
        entry = merged[date]
        ws.cell(row=i, column=1, value=date)
        ws.cell(row=i, column=2, value=entry["kg"])
        ws.cell(row=i, column=3, value=entry.get("notes") or None)


def seed(tracker_path: Path, entries: list[dict]) -> int:
    wb = openpyxl.load_workbook(tracker_path)
    ws = ensure_bodyweight_sheet(wb)

    merged = read_existing(ws)
    added = 0
    updated = 0
    for e in entries:
        n = normalize(e)
        if n["date"] in merged:
            if merged[n["date"]] != {"kg": n["kg"], "notes": n["notes"]}:
                updated += 1
            merged[n["date"]] = {"kg": n["kg"], "notes": n["notes"]}
        else:
            added += 1
            merged[n["date"]] = {"kg": n["kg"], "notes": n["notes"]}

    write_sorted(ws, merged)
    style_bodyweight_sheet(ws)
    wb.save(tracker_path)

    print(f"Seeded Bodyweight: {added} new, {updated} updated, {len(merged)} total")
    return 0


def load_entries(source: str) -> list[dict]:
    if source == "-":
        data = json.load(sys.stdin)
    else:
        data = json.loads(Path(source).read_text())
    if not isinstance(data, list):
        raise ValueError("entries JSON must be a list")
    for e in data:
        if "date" not in e or "kg" not in e:
            raise ValueError(f"entry missing date/kg: {e!r}")
    return data


def main() -> int:
    if len(sys.argv) != 3:
        print(__doc__, file=sys.stderr)
        return 2
    tracker = Path(sys.argv[1])
    if not tracker.exists():
        print(f"ERROR: tracker not found: {tracker}", file=sys.stderr)
        return 1
    entries = load_entries(sys.argv[2])
    return seed(tracker, entries)


if __name__ == "__main__":
    sys.exit(main())
