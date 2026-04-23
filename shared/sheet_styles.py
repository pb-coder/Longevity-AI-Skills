"""Canonical Workout Tracker styling.

Shared between /log (applied on every append) and /maintain (full-sheet
restyle). Idempotent: running `style_monthly_sheet` twice in a row is a no-op.

Keep this module the single source of truth for fonts, fills, borders,
column widths, and alignment.
"""
from datetime import datetime, date
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------ palette
HEADER_FILL = "FFBDC3C7"
DATA_FILL   = "FFF2F3F4"
NAVY        = "FF2C3E50"
LIGHT_GRAY  = "FFD5D8DC"
SUBSEC_FILL = "FFEAEDED"
BLACK = "FF000000"
WHITE = "FFFFFFFF"
SEP_COLOR = "FFBDC3C7"

font_header    = Font(bold=True, size=10, color=BLACK)
font_section_w = Font(bold=True, size=10, color=WHITE)
font_section_b = Font(bold=True, size=10, color=BLACK)
font_subsec    = Font(bold=True, italic=True, size=10, color=BLACK)
font_data      = Font(bold=False, size=10, color=BLACK)
font_total     = Font(bold=True, size=10, color=BLACK)

fill_header = PatternFill("solid", fgColor=HEADER_FILL)
fill_navy   = PatternFill("solid", fgColor=NAVY)
fill_gray   = PatternFill("solid", fgColor=LIGHT_GRAY)
fill_subsec = PatternFill("solid", fgColor=SUBSEC_FILL)
fill_data   = PatternFill("solid", fgColor=DATA_FILL)
fill_total  = PatternFill("solid", fgColor=LIGHT_GRAY)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
align_left   = Alignment(horizontal="left",   vertical="center", wrap_text=False)

sep_side = Side(style="thin", color=SEP_COLOR)
border_session = Border(top=sep_side)
no_border = Border()

# ------------------------------------------------------------------ structure
# Monthly sheet columns (A..M). A=SESSION (per-month number, merged per date).
# Notes column reads better left-aligned. Volume holds a formula, not a number.
MONTHLY_HEADERS = [
    "SESSION", "Date", "#", "Exercise", "Set", "Reps", "kg", "Volume", "Notes",
    "Distance (km)", "Duration (min)", "Pace (min/km)", "Avg HR",
]
MONTHLY_WIDTHS = {
    "A": 8,  "B": 12, "C": 5,  "D": 28, "E": 5,  "F": 6, "G": 6,
    "H": 9,  "I": 24, "J": 13, "K": 14, "L": 13, "M": 9,
}
MONTHLY_LEFT_COLS = {"I"}
MONTHLY_COLS = 13
TOTAL_LABEL = "TOTAL"

# Exercises Database: 5 columns (Exercise | Type | Primary Muscle | Equipment | Variations).
DB_WIDTHS = {"A": 30, "B": 13, "C": 16, "D": 14, "E": 48}
DB_COLS = 5

# Bodyweight sheet: 3 columns (Date | Kg | Notes). Morning / empty-stomach
# weigh-ins. Sorted DESC (newest at the top). Notes left-aligned; Date/Kg
# centered.
BODYWEIGHT_WIDTHS = {"A": 12, "B": 7, "C": 40}
BODYWEIGHT_LEFT_COLS = {"C"}
BODYWEIGHT_COLS = 3
BODYWEIGHT_HEADERS = ["Date", "Kg", "Notes"]


# ------------------------------------------------------------------ helpers
def find_last_data_cell(ws):
    """Return (last_data_row, last_data_col) — the max row/col with any value."""
    last_r, last_c = 0, 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and cell.value != "":
                if cell.row > last_r:
                    last_r = cell.row
                if cell.column > last_c:
                    last_c = cell.column
    return last_r, last_c


# ------------------------------------------------------------------ styling
def _to_num(v):
    """Coerce to float for strength-session detection. Blank/None → 0."""
    if v in (None, ""):
        return 0.0
    try:
        return float(str(v).replace(",", "."))
    except (TypeError, ValueError):
        return 0.0


def _date_str(v):
    """Coerce a Date cell value to a canonical ``YYYY-MM-DD`` string.

    Legacy rows imported via Numbers/Excel autoformat can land as
    ``datetime.datetime`` or ``datetime.date`` objects. The convention is
    strings, and mixing types breaks any comparison (sort, dict-key merge,
    equality-based session grouping). Normalise here so the rest of the
    styler can treat dates as strings. Unknown types pass through unchanged.
    """
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.isoformat()
    return v


def _numeric_cell(v):
    """Coerce stringified numbers (incl. European comma decimals like ``"67,5"``)
    to int or float. Returns the original value for anything that isn't purely
    numeric — MM:SS durations, blank cells, text notes — so callers can apply
    this indiscriminately to numeric columns without destroying non-numeric
    entries that happen to share a column historically.

    Excel requires real numbers in cells consumed by arithmetic formulas;
    ``=F*G`` on text returns #VALUE! (see the 2026-04-06 regression where
    legacy data imported kg as string ``"67,5"``). Normalising here keeps the
    sheet arithmetic-safe regardless of how the value was originally logged.
    """
    if v in (None, ""):
        return v
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if not s:
        return v
    try:
        f = float(s.replace(",", "."))
    except ValueError:
        return v
    return int(f) if f.is_integer() else f


def style_monthly_sheet(ws):
    """Apply canonical styling to a YYYY.MM workout log sheet.

    Idempotent. Running twice in a row is a no-op. Responsibilities:

    - Migrate legacy 12-col layout (A="Date") to the 13-col layout by
      inserting a leftmost SESSION column.
    - Normalise Date cells to ``YYYY-MM-DD`` strings (legacy datetime cells
      are coerced), then sort sessions by date ascending and merge any
      same-date sessions that became non-contiguous on disk — e.g. after a
      backfill row got appended at the bottom instead of into its day's
      block.
    - Rebuild SESSION column: per-month session numbers (1..N, chronological
      after the sort), merged vertically across each session's rows.
    - Rebuild TOTAL row at the end of each strength session with
      ``=SUM(H{first}:H{last})`` in the Volume column. Cardio-only sessions
      (every row has kg=0 and reps=0) get no TOTAL row.
    - Rewrite the Volume column on every set row as ``=F{r}*G{r}``
      (reps × kg) — the old hard-coded numeric volumes are discarded.
    - Apply header/data/TOTAL fills, fonts, widths, freeze pane.

    The data-row region is cleared and rebuilt from parsed rows so merges,
    TOTAL row placement, and session numbering can't drift out of sync with
    the data.
    """
    # ---- Migration: legacy layout had A="Date". Shift everything one col right.
    if ws.cell(row=1, column=1).value == "Date":
        ws.insert_cols(1)

    # ---- Unmerge any existing ranges so we can re-apply cleanly.
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))

    # ---- Pass 1: read all data rows into memory, grouped by contiguous date.
    last_r, _ = find_last_data_cell(ws)
    sessions: list[dict] = []
    current: dict | None = None
    for r in range(2, max(last_r, 1) + 1):
        date_val = _date_str(ws.cell(row=r, column=2).value)
        ex_val = ws.cell(row=r, column=4).value

        # Drop pre-existing TOTAL rows; we'll rebuild them.
        if ex_val == TOTAL_LABEL:
            continue
        # Drop fully-empty rows.
        if date_val in (None, "") and ex_val in (None, ""):
            continue

        row_data = {
            "date":      date_val,
            "num":       ws.cell(row=r, column=3).value,
            "exercise":  ex_val,
            "set":       ws.cell(row=r, column=5).value,
            "reps":      ws.cell(row=r, column=6).value,
            "kg":        ws.cell(row=r, column=7).value,
            "notes":     ws.cell(row=r, column=9).value,
            "distance":  ws.cell(row=r, column=10).value,
            "duration":  ws.cell(row=r, column=11).value,
            "pace":      ws.cell(row=r, column=12).value,
            "avg_hr":    ws.cell(row=r, column=13).value,
        }

        if current is None or date_val != current["date"]:
            current = {"date": date_val, "rows": []}
            sessions.append(current)
        current["rows"].append(row_data)

    # ---- Sort sessions by date ascending and merge any that share a date.
    # A shared-date pair happens when a backfill row (earlier than existing
    # data) got appended to the end of the sheet and is no longer contiguous
    # with that date's original block. Without this, backfilled workouts end
    # up stranded at the bottom and SESSION numbers drift out of chronological
    # order.
    merged: dict = {}
    for sess in sessions:
        date = sess["date"]
        merged.setdefault(date, {"date": date, "rows": []})["rows"].extend(sess["rows"])
    sessions = sorted(
        merged.values(),
        key=lambda s: (s["date"] is None, s["date"] or ""),
    )

    # ---- Clear data area.
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # ---- Pass 2: write header + data back with SESSION numbers, merges,
    # Volume formulas, TOTAL rows.
    for c, label in enumerate(MONTHLY_HEADERS, 1):
        ws.cell(row=1, column=c, value=label)

    write_row = 2
    for session_num, sess in enumerate(sessions, start=1):
        first_row = write_row

        # Strength session if any row contributes non-zero kg*reps volume.
        is_strength = any(
            _to_num(r["kg"]) * _to_num(r["reps"]) > 0 for r in sess["rows"]
        )

        for rd in sess["rows"]:
            ws.cell(row=write_row, column=1, value=session_num)
            ws.cell(row=write_row, column=2, value=rd["date"])
            ws.cell(row=write_row, column=3, value=_numeric_cell(rd["num"]))
            ws.cell(row=write_row, column=4, value=rd["exercise"])
            ws.cell(row=write_row, column=5, value=_numeric_cell(rd["set"]))
            ws.cell(row=write_row, column=6, value=_numeric_cell(rd["reps"]))
            ws.cell(row=write_row, column=7, value=_numeric_cell(rd["kg"]))
            ws.cell(row=write_row, column=8, value=f"=F{write_row}*G{write_row}")
            ws.cell(row=write_row, column=9, value=rd["notes"])
            ws.cell(row=write_row, column=10, value=_numeric_cell(rd["distance"]))
            ws.cell(row=write_row, column=11, value=rd["duration"])
            ws.cell(row=write_row, column=12, value=rd["pace"])
            ws.cell(row=write_row, column=13, value=_numeric_cell(rd["avg_hr"]))
            write_row += 1
        last_set_row = write_row - 1

        if is_strength:
            ws.cell(row=write_row, column=1, value=session_num)
            ws.cell(row=write_row, column=4, value=TOTAL_LABEL)
            ws.cell(row=write_row, column=8, value=f"=SUM(H{first_row}:H{last_set_row})")
            merge_last = write_row
            write_row += 1
        else:
            merge_last = last_set_row

        if merge_last > first_row:
            ws.merge_cells(
                start_row=first_row, end_row=merge_last,
                start_column=1, end_column=1,
            )

    last_written = write_row - 1

    # ---- Styling: header.
    for c in range(1, MONTHLY_COLS + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = no_border

    # ---- Styling: data rows + TOTAL rows. Skip merged follower cells —
    # openpyxl MergedCell objects don't accept style writes.
    for r in range(2, last_written + 1):
        is_total = ws.cell(row=r, column=4).value == TOTAL_LABEL
        for c in range(1, MONTHLY_COLS + 1):
            cell = ws.cell(row=r, column=c)
            if cell.__class__.__name__ == "MergedCell":
                continue
            cell.font = font_total if is_total else font_data
            cell.fill = fill_total if is_total else fill_data
            col = get_column_letter(c)
            cell.alignment = align_left if col in MONTHLY_LEFT_COLS else align_center
            cell.border = no_border

    # ---- Column widths + freeze pane.
    for col, w in MONTHLY_WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def style_bodyweight_sheet(ws):
    """Apply canonical styling to the Bodyweight sheet.

    Layout: Date | Kg | Notes. Data sorted DESC (newest on top) by the
    writers (seed_bodyweight.py, append_workout.upsert_bodyweight).
    Idempotent.

    Migrates the legacy 4-col layout (Year | Date | Kg | Notes) by dropping
    the leading Year column — Year was derivable from Date and added noise.
    """
    # Unmerge any existing merges (including legacy year merges) before edits.
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))

    # Migration: legacy layout had column A = "Year". Drop it.
    if ws.cell(row=1, column=1).value == "Year":
        ws.delete_cols(1)

    # Trim any stale column beyond the new layout (e.g. leftover col D="Notes"
    # after the Year delete shifts Notes to col C).
    if ws.max_column > BODYWEIGHT_COLS:
        ws.delete_cols(BODYWEIGHT_COLS + 1, ws.max_column - BODYWEIGHT_COLS)

    # Header row
    for c, label in enumerate(BODYWEIGHT_HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = no_border

    # Data rows: style every cell.
    last_data_row, _ = find_last_data_cell(ws)
    for r in range(2, last_data_row + 1):
        for c in range(1, BODYWEIGHT_COLS + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font_data
            cell.fill = fill_data
            col = get_column_letter(c)
            cell.alignment = align_left if col in BODYWEIGHT_LEFT_COLS else align_center
            cell.border = no_border

    # Column widths + freeze pane
    for col, w in BODYWEIGHT_WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"


def style_db_sheet(ws):
    """Reapply header + column widths on the Exercises Database.

    Does NOT reorder or rewrite body rows — assumes the DB is already organized
    by muscle > pattern. Only enforces: header row 1, col widths, freeze pane,
    and data-row fills for any bare rows.
    """
    # Header row
    headers = ["Exercise", "Type", "Primary Muscle", "Equipment", "Variations"]
    for c, label in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=label)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    # Restyle any data rows that lack the canonical fill.
    # Section headers (navy/gray) and subsections (EAEDED) keep their existing styles.
    last_data_row, _ = find_last_data_cell(ws)
    for r in range(2, last_data_row + 1):
        name = ws.cell(row=r, column=1).value
        typ  = ws.cell(row=r, column=2).value
        if not name or not typ:
            continue  # section header row — leave as is
        for c in range(1, DB_COLS + 1):
            cell = ws.cell(row=r, column=c)
            if cell.fill.fgColor is None or cell.fill.fgColor.rgb != DATA_FILL:
                cell.fill = fill_data
            if cell.font.size != 10 or cell.font.bold:
                cell.font = font_data
            if cell.alignment.horizontal != "center":
                cell.alignment = align_center

    # Widths + freeze
    for col, w in DB_WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
