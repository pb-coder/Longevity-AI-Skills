"""sync_db_sheet.py — Mirror exercises-database.md into the xlsx 'Exercises Database' tab.

Idempotent. For each section in the markdown, any entry that's missing in the
xlsx gets inserted at the end of that section. Sections that don't exist in
the xlsx yet are appended at the bottom with a navy section header. Existing
rows are never touched.

Usage:
    python3 sync_db_sheet.py "Workout Tracker - Nihad.xlsx"
"""
from __future__ import annotations
import sys
from pathlib import Path

import openpyxl

THIS_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(THIS_DIR))
from tracker_sheet import (  # noqa: E402
    style_db_sheet,
    fill_navy,
    font_section_w,
    align_center,
)

DB_MD = THIS_DIR / "exercises-database.md"
DB_SHEET = "Exercises Database"

EQUIP_MAP = {
    "BB": "Barbell", "DB": "Dumbbell", "Cable": "Cable", "Machine": "Machine",
    "BW": "Bodyweight", "Smith": "Smith", "LM": "Landmine", "Band": "Band",
    "Treadmill": "Treadmill", "Rower": "Rower", "Elliptical": "Elliptical",
    "Bike": "Bike", "Stair Climber": "Stair Climber",
    "Assault Bike": "Assault Bike", "Outdoor": "Outdoor", "Pool": "Pool",
    "Mat": "Mat", "Rope": "Jump Rope", "Foam Roller": "Foam Roller",
}


def parse_bullet(line: str) -> tuple[str, str | None] | None:
    """Return (name, equip_tag) or None if the line isn't an exercise bullet."""
    if not line.startswith("- "):
        return None
    s = line[2:].strip()
    if s.startswith("(") or ":" in s.split("[", 1)[0]:
        return None
    equip: str | None = None
    name = s
    if "[" in name:
        head, rest = name.split("[", 1)
        name = head.strip()
        if "]" in rest:
            equip = rest.split("]", 1)[0].strip()
    elif "—" in name:
        name = name.split("—", 1)[0].strip()
    return name.strip(), equip


def parse_md(md_path: Path) -> list[dict]:
    """Yield {section, subsection, name, equip} per bullet, in file order."""
    out: list[dict] = []
    section: str | None = None
    subsection: str | None = None
    for raw in md_path.read_text().splitlines():
        s = raw.rstrip()
        if s.startswith("## "):
            section = s[3:].strip()
            subsection = None
            continue
        if s.startswith("### "):
            subsection = s[4:].strip()
            continue
        parsed = parse_bullet(s)
        if not parsed or not section:
            continue
        name, equip = parsed
        out.append({
            "section": section,
            "subsection": subsection,
            "name": name,
            "equip": equip,
        })
    return out


def read_existing_names(ws) -> set[str]:
    names: set[str] = set()
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        typ = ws.cell(row=r, column=2).value
        if name and typ:
            names.add(str(name).strip().lower())
    return names


def find_section_rows(ws) -> dict[str, tuple[int, int]]:
    """Map UPPERCASE section name → (header_row, last_row_inclusive)."""
    headers: list[tuple[int, str]] = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        typ = ws.cell(row=r, column=2).value
        if not name or typ:
            continue
        font = ws.cell(row=r, column=1).font
        if font and font.bold and not font.italic:
            headers.append((r, str(name).strip()))
    out: dict[str, tuple[int, int]] = {}
    for i, (row, name) in enumerate(headers):
        last = headers[i + 1][0] - 1 if i + 1 < len(headers) else ws.max_row
        out[name.upper()] = (row, last)
    return out


def derive_type_and_muscle(entry: dict) -> tuple[str, str]:
    section = entry["section"].upper()
    sub = (entry["subsection"] or "").lower()
    if section == "WARMUP":
        return "Warmup", entry["subsection"] or "Full Body"
    if section == "CARDIO":
        return "Cardio", "Cardio"
    if section == "WELLNESS":
        return "Wellness", "Wellness"
    if "compound" in sub:
        return "Compound", section.title()
    return "Isolation", section.title()


def expand_equip(tag: str | None) -> str:
    if not tag:
        return ""
    return EQUIP_MAP.get(tag, tag)


def write_data_row(ws, row: int, entry: dict) -> None:
    typ, muscle = derive_type_and_muscle(entry)
    ws.cell(row=row, column=1, value=entry["name"])
    ws.cell(row=row, column=2, value=typ)
    ws.cell(row=row, column=3, value=muscle)
    ws.cell(row=row, column=4, value=expand_equip(entry["equip"]))


def write_section_header(ws, row: int, name: str) -> None:
    cell = ws.cell(row=row, column=1, value=name.upper())
    cell.font = font_section_w
    cell.fill = fill_navy
    cell.alignment = align_center


def main(xlsx_path: str) -> None:
    md_entries = parse_md(DB_MD)
    wb = openpyxl.load_workbook(xlsx_path)
    if DB_SHEET not in wb.sheetnames:
        raise SystemExit(f"{xlsx_path}: missing '{DB_SHEET}' sheet")
    ws = wb[DB_SHEET]

    existing = read_existing_names(ws)

    missing_by_section: dict[str, list[dict]] = {}
    section_order: list[str] = []
    for e in md_entries:
        if e["name"].lower() in existing:
            continue
        sec = e["section"].upper()
        if sec not in missing_by_section:
            missing_by_section[sec] = []
            section_order.append(sec)
        missing_by_section[sec].append(e)

    if not missing_by_section:
        print(f"{xlsx_path}: already in sync ({len(existing)} entries)")
        return

    additions = 0
    for sec in section_order:
        entries = missing_by_section[sec]
        section_rows = find_section_rows(ws)  # recompute after each block
        if sec in section_rows:
            _, last_row = section_rows[sec]
            insert_at = last_row + 1
            ws.insert_rows(insert_at, amount=len(entries))
            for i, e in enumerate(entries):
                write_data_row(ws, insert_at + i, e)
            print(f"  + {sec}: inserted {len(entries)} entries at row {insert_at}")
            additions += len(entries)
        else:
            bottom = ws.max_row + 1
            write_section_header(ws, bottom, sec)
            for i, e in enumerate(entries, start=1):
                write_data_row(ws, bottom + i, e)
            print(f"  + {sec}: appended new section at row {bottom} with {len(entries)} entries")
            additions += len(entries) + 1

    style_db_sheet(ws)
    wb.save(xlsx_path)
    print(f"{xlsx_path}: added {additions} rows; saved.")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("usage: sync_db_sheet.py <tracker.xlsx>", file=sys.stderr)
        sys.exit(2)
    main(sys.argv[1])
