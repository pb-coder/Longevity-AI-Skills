"""Read the given tracker xlsx for /coach analysis.

Emits one JSON blob on stdout organised around session-level signals
(decisions, not raw arrays). Top blocks:

  Source + capabilities:
  - today, data_source, capabilities, auto_cardio_enabled
  - estimated_max_hr, estimated_rest_hr — derived once at the top from
    Apple max-HR observations (XML) or 208 − 0.7×age (HL fallback;
    age is computed from Profile.birthday). Drives all HRR / TRIMP /
    Karvonen-zone math below.

  Strength + cardio sessions:
  - monthly_sessions: canonical per-session record incl. TRIMP /
    load_band (light/moderate/hard/red-line) / intensity_pct / max_hr /
    volume / is_deload, sourced from the TOTAL row's metadata + Apple
    per-workout max_hr. Replaces session_totals + workout_sessions_last_28d.
  - weekly_volume_per_muscle, estimated_1rm, progression_summary
  - stale_exercises (top 5), unknown_exercises
  - deloads (user-marked), auto_deload_candidates (Python-detected)

  Cardio rollup:
  - cardio_last_28d (sessions, minutes, distance, kcal)
  - cardio_hr_zones_28d (HRR-based time-in-zone via Karvonen)

  Recovery + training load (Python-derived):
  - recovery: 0-10 score with named drivers (HRV, RHR, sleep, wrist temp,
    HR Recovery 1-min, VO2max trend) + confidence
  - training_load: CTL/ATL/TSB rolling EWMA from per-session TRIMP
  - hr_at_volume_divergence: per-muscle slope of strength avg HR vs time

  Bodyweight:
  - bodyweight_latest, bodyweight_trend_kg_per_week

  Apple Health:
  - health_metrics_weekly (4-week aggregates; raw daily behind
    --include-daily-health)
  - vo2max_latest, vo2max_trend_per_4w

  Debug deep-dive (off by default):
  - rows: flat per-set list (--include-rows)
  - estimated_1rm.e1rm_history (--include-1rm-history)

Usage:
    python3 read_tracker.py "<tracker path>" [--months 3] [--today YYYY-MM-DD]
        [--include-rows] [--include-1rm-history] [--include-daily-health]
        [--pretty]

Keeping the model out of the weeds on format quirks (string vs datetime
dates, stringified numbers, casing inconsistency, empty-row streaks) is
the whole point — and going further, this script also computes the
training-science derivatives (TRIMP, recovery score, load bands, HR-at-
volume divergence) so the coach LLM consumes structured signals rather
than re-deriving them each run.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "shared"))
from tracker_sheet import (  # noqa: E402
    HEALTH_METRICS_COLS_BY_SOURCE,
    HEALTH_METRICS_FIELDS_BY_SOURCE,
    HEALTH_METRICS_SHEET_NAME,
    WORKOUT_SESSIONS_FIELDS_BY_SOURCE,
    WORKOUT_SESSIONS_SHEET_NAME,
    bw_locate_date,
    date_str,
    hm_locate_date,
    read_profile,
    ws_locate_date_start,
)

# Per-source capability map. The coach reads this to decide which sections of
# the report to write. ``xml`` is Apple's native zipped export (Nihad);
# ``hl_export`` is the HLExport text dump (Fabian) — much lighter, but no HRV,
# no wrist temp, no per-workout HR, no sleep stages, no Apple-aggregate RHR /
# walking HR / exercise-minute. The coach should distinguish ``unsupported``
# (data source can't provide it) from ``not yet collected`` (data source can,
# but the user hasn't logged enough yet).
SOURCE_CAPABILITIES = {
    "xml": {
        "hrv":                True,
        "wrist_temp":         True,
        "resting_hr_daily":   True,
        "walking_hr":         True,
        "sleep_stages":       True,
        "sleep_breath_dist":  True,
        "exercise_min_daily": True,
        "per_workout_hr":     True,
    },
    "hl_export": {
        "hrv":                False,
        "wrist_temp":         False,
        "resting_hr_daily":   False,
        "walking_hr":         False,
        "sleep_stages":       False,
        "sleep_breath_dist":  False,
        "exercise_min_daily": False,
        "per_workout_hr":     False,
    },
}

# Applied when the Profile sheet is missing or unset — treat the data as
# coming from XML so existing Nihad trackers (created before the Profile
# sheet existed) keep their full capability surface. New Fabian trackers
# get bootstrapped to ``hl_export`` by ``import_hl_export.py``.
DEFAULT_DATA_SOURCE = "xml"

MONTHLY_RE = re.compile(r"^\d{4}\.\d{2}$")
# Deload marker now lives on the TOTAL row's Notes column (col 9). The
# marker text is canonical "Deload Workout"; matching is case-insensitive.
DELOAD_MARKER = "deload workout"
EMPTY_STREAK_STOP = 10
TOTAL_LABEL = "TOTAL"

# Per-muscle weekly volume landmarks (hard sets). Source: references/training-
# science.md §1 + RP Strength tables. MV=maintenance, MEV=minimum effective,
# MAV=maximum adaptive, MRV=maximum recoverable. Numbers are individual and
# approximate; the coach uses them to name the band the current volume sits in.
VOLUME_LANDMARKS = {
    "chest":        {"mv": 6,  "mev": 10, "mav": 16, "mrv": 22},
    "back":         {"mv": 6,  "mev": 10, "mav": 18, "mrv": 25},
    "lats":         {"mv": 6,  "mev": 10, "mav": 16, "mrv": 22},
    "quads":        {"mv": 6,  "mev": 10, "mav": 18, "mrv": 22},
    "hamstrings":   {"mv": 4,  "mev": 8,  "mav": 14, "mrv": 18},
    "glutes":       {"mv": 4,  "mev": 8,  "mav": 14, "mrv": 18},
    "front_delts":  {"mv": 4,  "mev": 6,  "mav": 12, "mrv": 16},
    "side_delts":   {"mv": 6,  "mev": 8,  "mav": 16, "mrv": 22},
    "rear_delts":   {"mv": 6,  "mev": 8,  "mav": 16, "mrv": 22},
    "biceps":       {"mv": 5,  "mev": 8,  "mav": 14, "mrv": 20},
    "triceps":      {"mv": 4,  "mev": 6,  "mav": 12, "mrv": 18},
    "calves":       {"mv": 6,  "mev": 8,  "mav": 16, "mrv": 22},
    "forearms":     {"mv": 2,  "mev": 4,  "mav": 8,  "mrv": 12},
    "abs":          {"mv": 0,  "mev": 4,  "mav": 16, "mrv": 25},
    "core":         {"mv": 0,  "mev": 4,  "mav": 16, "mrv": 25},
    "erectors":     {"mv": 2,  "mev": 4,  "mav": 10, "mrv": 16},
    "traps":        {"mv": 2,  "mev": 4,  "mav": 10, "mrv": 16},
    "adductors":    {"mv": 0,  "mev": 2,  "mav": 8,  "mrv": 12},
    "neck":         {"mv": 0,  "mev": 2,  "mav": 6,  "mrv": 12},
}

# Canonicalise the muscle tokens that appear in exercises-database.md to the
# snake_case keys used everywhere else (and in VOLUME_LANDMARKS).
MUSCLE_ALIASES = {
    "chest": "chest", "upper chest": "upper_chest",
    "back": "back", "lats": "lats",
    "biceps": "biceps", "triceps": "triceps",
    "quads": "quads", "hamstrings": "hamstrings",
    "glutes": "glutes", "adductors": "adductors",
    "calves": "calves", "forearms": "forearms",
    "abs": "abs", "core": "core",
    "erectors": "erectors", "traps": "traps",
    "neck": "neck",
    "front delt": "front_delts", "front delts": "front_delts",
    "side delt": "side_delts",  "side delts":  "side_delts",
    "rear delt": "rear_delts",  "rear delts":  "rear_delts",
    "external rotators": "external_rotators",
    "shoulders": "shoulders",   "full body": "full_body",
    "posterior chain": None,    # too broad to assign — skip as primary
}

# Which ## SECTION header implies which primary muscle. None means "use
# subsection hint or parenthetical override". SHOULDERS is deliberately None
# because its subsections route to specific delt regions.
SECTION_PRIMARY = {
    "WARMUP": None, "CARDIO": None, "FULL BODY": None, "FULL BODY (COMPOUND)": None,
    "CHEST": "chest", "BACK": "back",
    "SHOULDERS": None,
    "BICEPS": "biceps", "TRICEPS": "triceps",
    "QUADS": "quads", "HAMSTRINGS": "hamstrings",
    "GLUTES": "glutes", "ADDUCTORS": "adductors",
    "CALVES": "calves", "CORE": "core",
    "NECK": "neck",
}

# Subsection hints that override the section heading (used inside SHOULDERS
# and for the stray "Forearms" subsection under BICEPS). Matched by substring
# against the lowercased subsection header.
SUBSECTION_PRIMARY_HINTS = [
    ("lateral delt", "side_delts"),
    ("rear delt",    "rear_delts"),
    ("vertical push","front_delts"),  # overhead press etc. primarily hit front delts
    ("traps",        "traps"),
    ("forearms",     "forearms"),
]


# ---------- helpers ----------
def to_float(v) -> float:
    if v in (None, ""):
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def to_int_or_none(v):
    if v in (None, ""):
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _compact(obj):
    """Recursively drop ``None``-valued keys from dicts.

    Applied once to the top-level payload before serialisation so every
    row, summary entry, and nested dict sheds its ``None`` ballast.
    ``0``, ``""``, ``False``, and empty collections are preserved —
    they carry meaning. An absent key and a key set to ``null`` are
    equivalent to an LLM reading the JSON as prose.
    """
    if isinstance(obj, dict):
        return {k: _compact(v) for k, v in obj.items() if v is not None}
    if isinstance(obj, list):
        return [_compact(v) for v in obj]
    return obj


def parse_duration_minutes(raw) -> float:
    """Accept '30:00', '28:30', '30', 30, 30.0 — return minutes as float."""
    if raw in (None, ""):
        return 0.0
    s = str(raw).strip()
    if ":" in s:
        parts = s.split(":")
        try:
            mm = int(parts[0])
            ss = int(parts[1]) if len(parts) > 1 else 0
            return mm + ss / 60.0
        except ValueError:
            return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_distance_km(raw) -> float:
    """Accept '5', '5.0', '8,79' (German decimal), 5, 5.0."""
    if raw in (None, ""):
        return 0.0
    s = str(raw).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


# ---------- extraction ----------
def extract_rows(wb, months_back: int, today_d: date) -> tuple[list[dict], dict, dict]:
    """Return (rows, session_totals, session_summaries).

    ``rows`` excludes TOTAL summary rows — one entry per logged set. Keys:
    ``session, date, num, exercise, set, reps, kg, volume, notes,
    distance_km, duration_min, pace, avg_hr, active_cal, total_cal,
    elevation_m, elapsed``.

    ``session_totals`` maps ``YYYY-MM-DD`` → total volume lifted that session,
    populated from the sheet's TOTAL rows (formula-driven).

    ``session_summaries`` maps ``YYYY-MM-DD`` → dict of session-level
    metadata harvested from the TOTAL row: ``volume`` (=session_totals
    value), ``notes`` (deload marker if present), ``duration_min``,
    ``avg_hr``, ``active_cal``, ``total_cal``, ``elevation_m``,
    ``elapsed``, ``is_deload`` (bool). Cardio-only dates (no TOTAL row)
    are absent from this dict.
    """
    cutoff = today_d - timedelta(days=months_back * 31)
    data_sheets = sorted(
        [s for s in wb.sheetnames if MONTHLY_RE.match(s)],
        reverse=True,
    )

    rows: list[dict] = []
    session_totals: dict[str, float] = {}
    session_summaries: dict[str, dict] = {}
    for name in data_sheets:
        # Quick filter: sheet YYYY.MM vs cutoff
        y, m = name.split(".")
        first_of_month = date(int(y), int(m), 1)
        if first_of_month < cutoff.replace(day=1):
            continue

        ws = wb[name]
        current_date: str | None = None
        empty_streak = 0

        for raw in ws.iter_rows(min_row=2, values_only=True):
            padded = list(raw) + [None] * 17
            (session, date_val, num, exercise, set_n, reps, kg, volume,
             notes, distance, duration, pace, avg_hr,
             active_cal, total_cal, elevation_m, elapsed) = padded[:17]

            if date_val is None and exercise is None:
                empty_streak += 1
                if empty_streak >= EMPTY_STREAK_STOP:
                    break
                continue
            empty_streak = 0

            if date_val is not None:
                current_date = date_str(date_val)

            # TOTAL rows now carry the session's full summary record:
            # Date (col 2 may be set), Volume (col 8), Notes (col 9 —
            # carries the Deload Workout marker), Duration (col 11),
            # Avg HR (col 13), Active/Total Cal (cols 14-15), Elevation
            # (col 16), Elapsed (col 17). Harvest those into
            # session_summaries keyed by date.
            if isinstance(exercise, str) and exercise.strip().upper() == TOTAL_LABEL:
                # Prefer the TOTAL row's own Date if set; fall back to
                # current_date from the preceding non-TOTAL rows for
                # legacy data.
                total_date = current_date
                if date_val is not None:
                    parsed = date_str(date_val)
                    if parsed:
                        total_date = parsed
                if total_date is not None:
                    if volume not in (None, ""):
                        session_totals[total_date] = to_float(volume)
                    notes_str = str(notes).strip() if notes else None
                    is_deload = bool(notes_str and DELOAD_MARKER in notes_str.lower())
                    session_summaries[total_date] = {
                        "volume":       session_totals.get(total_date),
                        "notes":        notes_str,
                        "is_deload":    is_deload,
                        "duration_min": parse_duration_minutes(duration) if duration else None,
                        "avg_hr":       to_float(avg_hr) if avg_hr is not None else None,
                        "active_cal":   to_float(active_cal) if active_cal not in (None, "") else None,
                        "total_cal":    to_float(total_cal) if total_cal not in (None, "") else None,
                        "elevation_m":  to_float(elevation_m) if elevation_m not in (None, "") else None,
                        "elapsed":      str(elapsed).strip() if elapsed not in (None, "") else None,
                    }
                continue

            if exercise is None or current_date is None:
                continue

            # Volume is formula-driven in the sheet; if Excel hasn't cached it,
            # recompute from kg × reps so downstream consumers never see None.
            reps_i = to_int_or_none(reps)
            kg_f = to_float(kg)
            if volume in (None, ""):
                vol_f = kg_f * (reps_i or 0)
            else:
                vol_f = to_float(volume)

            rows.append({
                "session": to_int_or_none(session),
                "date": current_date,
                "num": num,
                "exercise": str(exercise).strip(),
                "set": set_n,
                "reps": reps_i,
                "kg": kg_f,
                "volume": vol_f,
                "notes": (str(notes).strip() if notes else None),
                "distance_km": parse_distance_km(distance) if distance else None,
                "duration_min": parse_duration_minutes(duration) if duration else None,
                "pace": str(pace).strip() if pace else None,
                "avg_hr": to_int_or_none(avg_hr),
                "active_cal":  to_float(active_cal) if active_cal not in (None, "") else None,
                "total_cal":   to_float(total_cal) if total_cal not in (None, "") else None,
                "elevation_m": to_float(elevation_m) if elevation_m not in (None, "") else None,
                "elapsed":     str(elapsed).strip() if elapsed not in (None, "") else None,
            })

    rows.sort(key=lambda r: (r["date"], r["num"] or 0, r["set"] or 0))

    # Fill in session_totals for any date whose TOTAL cell lacked a cached
    # value (common when openpyxl reads formulas Excel hasn't saved yet).
    # Trust the sheet first — only sum rows for dates the sheet didn't cover.
    cached_dates = set(session_totals.keys())
    for r in rows:
        if r["date"] in cached_dates:
            continue
        session_totals[r["date"]] = session_totals.get(r["date"], 0.0) + r["volume"]

    return rows, session_totals, session_summaries


def progression_summary(rows: list[dict]) -> list[dict]:
    """Last and previous best working set per exercise (warmups excluded)."""
    by_ex: dict[str, list[dict]] = {}
    for r in rows:
        if r.get("notes") and "warmup" in r["notes"].lower():
            continue
        if not r.get("kg") or not r.get("reps"):
            continue
        by_ex.setdefault(r["exercise"].lower(), []).append(r)

    summary = []
    for canon_lower, sets in by_ex.items():
        # Group by date, pick heaviest (kg, then reps) per session.
        by_date: dict[str, dict] = {}
        for s in sets:
            cur = by_date.get(s["date"])
            if cur is None or (s["kg"], s["reps"]) > (cur["kg"], cur["reps"]):
                by_date[s["date"]] = s
        dates_desc = sorted(by_date.keys(), reverse=True)
        if len(dates_desc) < 1:
            continue
        last = by_date[dates_desc[0]]
        prev = by_date[dates_desc[1]] if len(dates_desc) >= 2 else None
        summary.append({
            "exercise": last["exercise"],
            "sessions_logged": len(dates_desc),
            "last": f"{dates_desc[0]} → {int(last['kg'])}kg x {last['reps']}",
            "prev": f"{dates_desc[1]} → {int(prev['kg'])}kg x {prev['reps']}" if prev else None,
        })

    summary.sort(key=lambda s: s["exercise"].lower())
    return summary


def find_deloads(wb) -> list[str]:
    """Dates whose strength session's TOTAL row has Notes containing 'Deload Workout'.

    Deload marker now lives on the TOTAL row's Notes column (col 9),
    consistent with the other session-level metadata. The TOTAL row's
    Date (col 2) is the canonical date; fall back to ``current_date``
    from preceding rows for legacy data.
    """
    deloads: set[str] = set()
    for name in wb.sheetnames:
        if not MONTHLY_RE.match(name):
            continue
        ws = wb[name]
        current_date: str | None = None
        empty_streak = 0
        for raw in ws.iter_rows(min_row=2, values_only=True):
            vals = list(raw) + [None] * 13
            _session, date_val, _num, exercise, _set_n, _reps, _kg, _vol, notes = vals[:9]
            if date_val is None and exercise is None:
                empty_streak += 1
                if empty_streak >= EMPTY_STREAK_STOP:
                    break
                continue
            empty_streak = 0
            if date_val is not None:
                parsed = date_str(date_val)
                if parsed:
                    current_date = parsed
            if exercise is None:
                continue
            if isinstance(exercise, str) and exercise.strip().upper() == TOTAL_LABEL:
                target_date = current_date
                if date_val is not None:
                    parsed = date_str(date_val)
                    if parsed:
                        target_date = parsed
                if target_date and notes and DELOAD_MARKER in str(notes).lower():
                    deloads.add(target_date)
    return sorted(deloads)


def read_bodyweight(wb) -> list[dict]:
    """Return all Bodyweight entries sorted ascending by date.

    Each entry: {"date": "YYYY-MM-DD", "kg": float, "notes": str|None}.
    Returns [] if the sheet is missing. The sheet stores newest-first with
    a per-year merge on column A, but this function re-sorts ascending so
    the trend/recent helpers see a stable chronological order.
    """
    if "Bodyweight" not in wb.sheetnames:
        return []
    ws = wb["Bodyweight"]
    out: list[dict] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not raw:
            continue
        d, date_idx = bw_locate_date(raw)
        if d is None:
            continue
        kg_raw = raw[date_idx + 1] if len(raw) > date_idx + 1 else None
        try:
            kg = float(kg_raw) if kg_raw not in (None, "") else None
        except (TypeError, ValueError):
            continue
        if kg is None:
            continue
        notes = raw[date_idx + 2] if len(raw) > date_idx + 2 else None
        out.append({
            "date": d,
            "kg": kg,
            "notes": (str(notes).strip() if notes else None),
        })
    out.sort(key=lambda e: e["date"])
    return out


def bodyweight_trend_kg_per_week(entries: list[dict]) -> float | None:
    """Simple slope over the last 8 entries: (last_kg - first_kg) / weeks_between.

    Returns None if fewer than 3 entries or the span is <7 days (too noisy).
    Excludes entries with notes flagging non-morning/non-fasted context.
    """
    clean = [e for e in entries if not _is_flagged_nonfasted(e)]
    window = clean[-8:]
    if len(window) < 3:
        return None
    first_d = datetime.strptime(window[0]["date"], "%Y-%m-%d").date()
    last_d = datetime.strptime(window[-1]["date"], "%Y-%m-%d").date()
    days = (last_d - first_d).days
    if days < 7:
        return None
    weeks = days / 7.0
    return round((window[-1]["kg"] - window[0]["kg"]) / weeks, 3)


def _is_flagged_nonfasted(entry: dict) -> bool:
    notes = (entry.get("notes") or "").lower()
    return any(k in notes for k in ("not fasted", "evening", "after", "post-meal"))


def build_monthly_sessions(rows: list[dict],
                            session_summaries: dict[str, dict] | None = None,
                            session_totals: dict[str, float] | None = None,
                            apple_sessions: list[dict] | None = None,
                            ) -> list[dict]:
    """Aggregate per-set rows into one entry per session-date.

    Strength sessions: metadata sourced from the TOTAL row's summary
    record in ``session_summaries`` (Active Cal, Total Cal, Elevation,
    Elapsed, Avg HR, Duration). Cardio-only sessions don't have a TOTAL
    row, so their metadata is read directly from the cardio rows.

    Folds in:
    - ``volume`` for strength sessions from ``session_totals`` (so the
      caller doesn't need to ship ``session_totals`` separately).
    - ``max_hr`` per session from ``apple_sessions`` (Apple's per-workout
      max HR — only present for XML; HL surfaces None and the field is
      stripped by ``_compact``).

    Returns a list sorted by date ascending.
    """
    summaries = session_summaries or {}
    totals = session_totals or {}
    apple = apple_sessions or []

    # date → max_hr lookup. Apple may record multiple workouts per date
    # (Core + Functional + cardio rides); we keep the largest max_hr seen
    # across all of them as the session's peak. We deliberately don't
    # surface ``apple_type`` here because it conflates strength and
    # cardio for mixed days — ``session_kind`` is the authoritative tag.
    by_date_apple: dict[str, float] = {}
    for ap in apple:
        d = ap.get("date")
        if not d:
            continue
        mh = ap.get("max_hr")
        if mh and mh > by_date_apple.get(d, 0):
            by_date_apple[d] = mh

    by_date: dict[str, dict] = {}
    for r in rows:
        d = r.get("date")
        if not d:
            continue
        is_strength_row = (r.get("kg") or 0) * (r.get("reps") or 0) > 0
        is_cardio_row = (r.get("distance_km") or 0) > 0 or (r.get("duration_min") or 0) > 0

        s = by_date.get(d)
        if s is None:
            s = {
                "date": d,
                "exercise_first": r.get("exercise"),
                "active_cal":  None,
                "total_cal":   None,
                "elevation_m": None,
                "elapsed":     None,
                "avg_hr":      None,
                "duration_min": None,
                "_has_strength": is_strength_row,
                "_has_cardio":   is_cardio_row,
            }
            by_date[d] = s
        else:
            if is_strength_row:
                s["_has_strength"] = True
            if is_cardio_row:
                s["_has_cardio"] = True

        # For cardio-only sessions: fill metadata from each cardio row.
        # For mixed/strength sessions, the TOTAL row summary is canonical
        # and is folded in below.
        if is_cardio_row and not is_strength_row:
            for k in ("active_cal", "total_cal", "elevation_m", "elapsed", "avg_hr"):
                if s.get(k) in (None, "") and r.get(k) not in (None, ""):
                    s[k] = r.get(k)
            if s.get("duration_min") in (None, "") and r.get("duration_min"):
                s["duration_min"] = r.get("duration_min")

    # Fold TOTAL-row session summaries (strength sessions only — TOTAL
    # rows are not emitted for pure cardio).
    for d, summary in summaries.items():
        s = by_date.get(d)
        if s is None:
            continue
        if summary.get("active_cal") is not None:
            s["active_cal"] = summary["active_cal"]
        if summary.get("total_cal") is not None:
            s["total_cal"] = summary["total_cal"]
        if summary.get("elevation_m") is not None:
            s["elevation_m"] = summary["elevation_m"]
        if summary.get("elapsed"):
            s["elapsed"] = summary["elapsed"]
        if summary.get("avg_hr") is not None:
            s["avg_hr"] = summary["avg_hr"]
        if summary.get("duration_min") is not None:
            s["duration_min"] = summary["duration_min"]
        if summary.get("is_deload"):
            s["is_deload"] = True

    out: list[dict] = []
    for d in sorted(by_date.keys()):
        s = by_date[d]
        kind = "strength" if s.pop("_has_strength") else (
            "cardio" if s.pop("_has_cardio") else "other")
        s.pop("_has_cardio", None)
        s["session_kind"] = kind
        # Fold in volume (strength only) and Apple max_hr.
        if kind == "strength" and d in totals:
            s["volume"] = totals[d]
        max_hr = by_date_apple.get(d)
        if max_hr:
            s["max_hr"] = max_hr
        out.append(s)
    return out


def cardio_last_28d(rows: list[dict], today_d: date) -> dict:
    """4-week cardio rollup: total distance, total minutes, total cal, and
    a coarse intervals-vs-zone2 split.

    Now uses a 28d window (was 14d) to align with the strength-side
    weekly_volume window. Cardio rows are identified by distance or
    duration > 0; intervals are flagged from Notes keywords or avg_hr
    >= 165 (a rough ceiling that catches Z4+ work without the user
    having to annotate).
    """
    cutoff = today_d - timedelta(days=28)
    zone2_min = 0.0
    intervals = 0
    distance = 0.0
    total_min = 0.0
    total_cal = 0.0
    sessions = 0
    for r in rows:
        try:
            d = datetime.strptime(r["date"], "%Y-%m-%d").date()
        except ValueError:
            continue
        if d < cutoff:
            continue
        dur = r.get("duration_min") or 0
        dist = r.get("distance_km") or 0
        if dur == 0 and dist == 0:
            continue
        sessions += 1
        distance += dist
        total_min += dur
        cal = r.get("active_cal") or 0
        total_cal += cal
        note = (r.get("notes") or "").lower()
        hr = r.get("avg_hr") or 0
        is_intervals = any(k in note for k in ("interval", "zone 4", "zone 5", "z4", "z5")) or hr >= 165
        if is_intervals:
            intervals += 1
        else:
            zone2_min += dur
    return {
        "sessions":          sessions,
        "total_minutes":     round(total_min, 1),
        "total_distance_km": round(distance, 2),
        "total_active_cal":  int(round(total_cal)) if total_cal else 0,
        "zone2_minutes":     round(zone2_min, 1),
        "interval_sessions": intervals,
    }


# ---------- health metrics ----------
def read_health_metrics(wb) -> list[dict]:
    """Return all Health Metrics rows sorted ascending by date.

    Each entry: ``{"date": "YYYY-MM-DD", <field>: <value>|None, ...}``,
    with one key per ``HEALTH_METRICS_FIELDS`` plus ``notes`` from the
    sheet's manual Notes column. Returns ``[]`` if the sheet is missing
    or empty.

    The sheet stores newest-first (the importer writes DESC); this
    function re-sorts ascending so trend/rolling helpers see a stable
    chronological order.
    """
    if HEALTH_METRICS_SHEET_NAME not in wb.sheetnames:
        return []
    src = read_profile(wb).get("source") or "xml"
    if src not in HEALTH_METRICS_FIELDS_BY_SOURCE:
        src = "xml"
    fields = HEALTH_METRICS_FIELDS_BY_SOURCE[src]
    notes_idx = HEALTH_METRICS_COLS_BY_SOURCE[src] - 1  # zero-based notes col
    # Keys callers may legitimately query but that the active source can't
    # populate (HL slim schema). Surface them as None so downstream
    # capability gates and trend helpers don't KeyError.
    all_xml_keys = HEALTH_METRICS_FIELDS_BY_SOURCE["xml"]
    missing_keys = [k for k in all_xml_keys if k not in fields]

    ws = wb[HEALTH_METRICS_SHEET_NAME]
    out: list[dict] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not raw:
            continue
        d, _ = hm_locate_date(raw)
        if d is None:
            continue
        entry = {"date": d}
        for i, key in enumerate(fields, start=1):
            v = raw[i] if len(raw) > i else None
            if v in (None, ""):
                entry[key] = None
            else:
                try:
                    entry[key] = float(v)
                except (TypeError, ValueError):
                    entry[key] = None
        for k in missing_keys:
            entry[k] = None
        notes = raw[notes_idx] if len(raw) > notes_idx else None
        entry["notes"] = (str(notes).strip() if notes else None)
        out.append(entry)
    out.sort(key=lambda e: e["date"])
    return out


def read_workout_sessions(wb) -> list[dict]:
    """Return all Workout Sessions rows sorted ascending by date+start.

    Each entry has the per-source columns of the sheet (xml: 12 cols incl.
    Avg/Max/Min HR; hl_export: 9 cols, HR fields surfaced as None).
    Returns ``[]`` if the sheet is missing.
    """
    if WORKOUT_SESSIONS_SHEET_NAME not in wb.sheetnames:
        return []
    src = read_profile(wb).get("source") or "xml"
    if src not in WORKOUT_SESSIONS_FIELDS_BY_SOURCE:
        src = "xml"
    fields = WORKOUT_SESSIONS_FIELDS_BY_SOURCE[src]

    # Per-key coercer. Keys absent from the active source's field list
    # (HL: avg/max/min HR) get None.
    numeric_keys = {"duration_min", "avg_hr", "active_cal", "distance_km"}
    int_keys     = {"max_hr", "min_hr"}

    ws = wb[WORKOUT_SESSIONS_SHEET_NAME]
    out: list[dict] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        d, s = ws_locate_date_start(raw)
        if d is None:
            continue
        entry = {"date": d, "start": s}
        # Map fields[i] → raw[i+1] (col 1 is Date, col 2..N is fields[0..N-1]).
        for i, key in enumerate(fields, start=1):
            v = raw[i] if len(raw) > i else None
            if key in numeric_keys:
                entry[key] = to_float(v) if v is not None else 0.0
            elif key in int_keys:
                entry[key] = to_int_or_none(v) if v is not None else None
            elif key == "notes":
                entry[key] = str(v).strip() if v else None
            else:
                entry[key] = v
        # Backfill keys missing from the active schema so downstream code
        # (e.g. session-HR cross-check) doesn't KeyError on HL trackers.
        for missing in ("avg_hr", "max_hr", "min_hr",
                        "duration_min", "active_cal", "distance_km",
                        "end", "apple_type", "source", "notes"):
            entry.setdefault(missing,
                             0.0 if missing in numeric_keys else None)
        out.append(entry)
    out.sort(key=lambda e: (e["date"], e["start"] or ""))
    return out


def _values_in_window(entries: list[dict], key: str, today_d: date, days: int) -> list[float]:
    cutoff = today_d - timedelta(days=days)
    out = []
    for e in entries:
        v = e.get(key)
        if v is None:
            continue
        try:
            d = datetime.strptime(e["date"], "%Y-%m-%d").date()
        except ValueError:
            continue
        if d < cutoff:
            continue
        out.append(float(v))
    return out


def _mean_or_none(values: list[float]) -> float | None:
    if not values:
        return None
    return round(sum(values) / len(values), 2)


def metric_trend_per_4w(entries: list[dict], key: str) -> float | None:
    """OLS slope of ``key`` over time, scaled to a 4-week window.

    Returns None unless there are ≥4 non-null entries spanning ≥21 days.
    Used for HRV / RHR / VO2max / sleep trends. Negative is improving for
    RHR (lower resting HR = better cardio fitness); positive is improving
    for HRV and VO2max.
    """
    pts: list[tuple[date, float]] = []
    for e in entries:
        v = e.get(key)
        if v is None:
            continue
        try:
            d = datetime.strptime(e["date"], "%Y-%m-%d").date()
        except ValueError:
            continue
        pts.append((d, float(v)))
    if len(pts) < 4:
        return None
    pts.sort(key=lambda p: p[0])
    span_days = (pts[-1][0] - pts[0][0]).days
    if span_days < 21:
        return None
    base = pts[0][0]
    xs = [(p[0] - base).days for p in pts]
    ys = [p[1] for p in pts]
    n = len(xs)
    mx = sum(xs) / n
    my = sum(ys) / n
    num = sum((xs[i] - mx) * (ys[i] - my) for i in range(n))
    den = sum((xs[i] - mx) ** 2 for i in range(n))
    if den <= 0:
        return None
    return round((num / den) * 28.0, 2)


def latest_metric(entries: list[dict], key: str) -> dict | None:
    """Most recent (date, value) pair for ``key``, or None if absent."""
    for e in reversed(entries):
        v = e.get(key)
        if v is not None:
            return {"date": e["date"], "value": round(float(v), 2)}
    return None


def baseline_60d(entries: list[dict], key: str, today_d: date) -> float | None:
    """Mean of ``key`` over the last 60 days. Used as anomaly baseline."""
    return _mean_or_none(_values_in_window(entries, key, today_d, 60))


def workout_sessions_in_window(sessions: list[dict], today_d: date, days: int) -> list[dict]:
    cutoff = today_d - timedelta(days=days)
    out = []
    for s in sessions:
        try:
            d = datetime.strptime(s["date"], "%Y-%m-%d").date()
        except ValueError:
            continue
        if d < cutoff:
            continue
        # Filter out incidental walks — those aren't training.
        if (s.get("notes") or "").lower().startswith("incidental"):
            continue
        out.append({
            "date":         s["date"],
            "type":         s.get("apple_type"),
            "duration":     s.get("duration_min"),
            "avg_hr":       s.get("avg_hr"),
            "max_hr":       s.get("max_hr"),
            "cal":          s.get("active_cal"),
        })
    return out


def strength_session_avg_hr_trend(
    sessions: list[dict],
    strength_dates: set[str],
) -> float | None:
    """Slope per 4 weeks of avg HR over the last 8 strength sessions.

    Matches Apple workouts to logged strength sessions by date. A rising
    avg HR on a stable load is a fatigue signal; the planning rule uses
    this to hold load when HR is creeping up.
    """
    matched: list[tuple[date, float]] = []
    for s in sessions:
        if s.get("date") not in strength_dates:
            continue
        avg = s.get("avg_hr")
        if avg in (None, 0):
            continue
        try:
            d = datetime.strptime(s["date"], "%Y-%m-%d").date()
        except ValueError:
            continue
        matched.append((d, float(avg)))
    if len(matched) < 4:
        return None
    matched.sort(key=lambda p: p[0])
    matched = matched[-8:]
    span_days = (matched[-1][0] - matched[0][0]).days
    if span_days < 21:
        return None
    base = matched[0][0]
    xs = [(p[0] - base).days for p in matched]
    ys = [p[1] for p in matched]
    n = len(xs)
    mx = sum(xs) / n
    my = sum(ys) / n
    num = sum((xs[i] - mx) * (ys[i] - my) for i in range(n))
    den = sum((xs[i] - mx) ** 2 for i in range(n))
    if den <= 0:
        return None
    return round((num / den) * 28.0, 2)


# ---------- exercises database ----------
_BULLET_RE = re.compile(
    r"^\s*-\s+"
    r"(?P<name>.+?)"                                 # exercise name
    r"(?:\s+\[(?P<equip>[^\]]+)\])?"                # [EQUIPMENT] optional
    r"(?:\s*—\s*(?P<syn>[^◆(]+?))?"                 # synergist list after em-dash
    r"(?P<leng>\s*◆)?"                              # optional lengthened flag
    r"(?:\s*\((?P<note>[^)]+)\))?"                  # optional parenthetical note
    r"\s*$"
)


def _canon_muscle(tok: str) -> str | None:
    """Map a raw muscle token from the database to a canonical key (or None)."""
    t = tok.strip().lower()
    # Strip trailing "s" pluralisation only for a short whitelist — avoid
    # collapsing "lats" and "hamstrings" which are already plural canonical.
    if t in MUSCLE_ALIASES:
        return MUSCLE_ALIASES[t]
    # Secondary pass: split on "/" for things like "erectors/lower back".
    for part in re.split(r"[/,]", t):
        p = part.strip()
        if p in MUSCLE_ALIASES:
            return MUSCLE_ALIASES[p]
    return None


def _primary_from_note(note: str | None) -> str | None:
    """If the bullet has ``(primary: X)``, return X's canonical muscle."""
    if not note:
        return None
    m = re.match(r"\s*primary:\s*(.+)$", note.strip(), re.IGNORECASE)
    if not m:
        return None
    return _canon_muscle(m.group(1))


def load_exercises_db(path: Path) -> dict[str, dict]:
    """Parse ``exercises-database.md`` into a dict keyed by lowercased
    exercise name. Each value:
      ``{"primary": str|None, "synergists": [str], "lengthened": bool,
         "equipment": str|None, "is_warmup": bool}``

    Unknown muscle tokens are silently dropped (not mapped) — the caller
    can detect gaps by comparing logged exercise names against this dict
    and surfacing anything missing as ``unknown_exercises``.
    """
    db: dict[str, dict] = {}
    if not path.exists():
        return db

    section = None        # e.g. "CHEST"
    subsection_primary = None  # regional override from a ### hint
    section_primary = None     # derived from SECTION_PRIMARY[section]

    for line in path.read_text().splitlines():
        s = line.rstrip()
        if s.startswith("## "):
            section = s[3:].strip().upper()
            section_primary = SECTION_PRIMARY.get(section)
            subsection_primary = None
            continue
        if s.startswith("### "):
            sub = s[4:].strip().lower()
            subsection_primary = None
            for key, muscle in SUBSECTION_PRIMARY_HINTS:
                if key in sub:
                    subsection_primary = muscle
                    break
            continue
        m = _BULLET_RE.match(s)
        if not m:
            continue

        name = m.group("name").strip()
        # Bullets in prose paragraphs (e.g. "(Biceps receive ~0.5 sets from…)")
        # are parenthetical, not exercises.
        if name.startswith("(") or ":" in name:
            continue

        equip = (m.group("equip") or "").strip() or None
        lengthened = m.group("leng") is not None
        note = m.group("note")
        raw_syn = m.group("syn") or ""

        synergists: list[str] = []
        for tok in raw_syn.split(","):
            tok = tok.strip()
            if not tok:
                continue
            # Each synergist in the database is written as "+muscle".
            if tok.startswith("+"):
                tok = tok[1:].strip()
            canon = _canon_muscle(tok)
            if canon:
                synergists.append(canon)

        # Primary resolution order: (primary: X) override → subsection hint
        # → section heading. None falls through for untagged sections (WARMUP,
        # CARDIO, FULL BODY) — those exercises get zero volume attribution.
        primary = (
            _primary_from_note(note)
            or subsection_primary
            or section_primary
        )

        db[name.lower()] = {
            "primary": primary,
            "synergists": synergists,
            "lengthened": lengthened,
            "equipment": equip,
            "is_warmup": section == "WARMUP",
            "is_cardio": section == "CARDIO",
        }
    return db


# ---------- derived coach features ----------
def _is_working_set(r: dict) -> bool:
    """A working set has a positive rep count and no 'warmup' in Notes.
    Bodyweight sets (kg=0, reps>0 like Pull-Up or Plank) count. Cardio rows
    (reps=0) and warmup-tagged rows are skipped."""
    reps = r.get("reps") or 0
    if reps <= 0:
        return False
    notes = (r.get("notes") or "").lower()
    if "warmup" in notes:
        return False
    return True


def weekly_volume_per_muscle(
    rows: list[dict],
    db: dict[str, dict],
    today_d: date,
    window_days: int,
    unknown_out: set[str],
) -> dict:
    """Fractional hard-set count per muscle over the last ``window_days``.

    Primary muscle = 1.0 set, each synergist = 0.5 set (per training-science
    §1). Warmup exercises (database section) and warmup-marked sets are
    skipped. Unknown exercises — logged names that don't appear in the db —
    are collected into ``unknown_out`` for the caller to surface.
    """
    cutoff = today_d - timedelta(days=window_days)
    sets: dict[str, float] = defaultdict(float)
    for r in rows:
        if not _is_working_set(r):
            continue
        try:
            d = datetime.strptime(r["date"], "%Y-%m-%d").date()
        except ValueError:
            continue
        if d < cutoff:
            continue
        entry = db.get(r["exercise"].lower())
        if entry is None:
            unknown_out.add(r["exercise"])
            continue
        if entry.get("is_warmup"):
            continue
        if entry["primary"]:
            sets[entry["primary"]] += 1.0
        for syn in entry["synergists"]:
            sets[syn] += 0.5

    current = {m: round(v, 1) for m, v in sets.items()}
    landmarks = {m: VOLUME_LANDMARKS[m] for m in current if m in VOLUME_LANDMARKS}
    return {
        "window_days": window_days,
        "current": current,
        "landmarks": landmarks,
    }


# History length cap for ``estimated_1rm[exercise].e1rm_history``. The slope
# field already summarises the trajectory; the LLM rarely needs more than the
# top of the list to spot-check confidence and grain. Cutting from 6 → 3
# entries removes ~50% of the per-exercise payload.
E1RM_HISTORY_LIMIT = 3


def estimated_1rm(rows: list[dict],
                  deload_dates: list[str] | None = None,
                  include_history: bool = False) -> dict:
    """Epley 1RM projection per exercise, with trajectory and confidence.

    For each exercise, take the heaviest projected e1RM per date (over all
    working sets that session) and report:
      - current/prev/best/last_date and current-vs-prev delta in kg
      - e1rm_history: last 6 sessions newest-first, each with the top set
        that produced the e1RM (so the coach can judge rep-range quality)
      - slope_kg_per_4w: OLS slope over the last 6 sessions, scaled to a
        4-week window. Null if fewer than 3 sessions.
      - confidence: high|medium|low based on the rep ranges of the last
        3 top sets — Epley is most accurate at 3-8 reps.
      - stalled_sessions: count of consecutive most-recent sessions with
        |Δe1RM| ≤ 0.5kg, broken by any deload that falls in the window.

    Bodyweight and warmup sets excluded (kg must be > 0).
    """
    deload_set = set(deload_dates or [])

    by_ex: dict[str, list[dict]] = {}
    canonical_name: dict[str, str] = {}
    for r in rows:
        if not _is_working_set(r):
            continue
        kg = r.get("kg") or 0
        reps = r.get("reps") or 0
        if kg <= 0 or reps <= 0:
            continue
        key = r["exercise"].lower()
        canonical_name.setdefault(key, r["exercise"])
        e1rm = kg * (1.0 + reps / 30.0)
        by_ex.setdefault(key, []).append({
            "date": r["date"], "e1rm": e1rm, "reps": reps, "kg": kg,
        })

    out: dict[str, dict] = {}
    for key, entries in by_ex.items():
        # Per date, keep the heaviest projected e1RM and remember the
        # (reps, kg) that produced it — needed for the history block and
        # for the confidence judgement.
        per_date: dict[str, dict] = {}
        for e in entries:
            top = per_date.get(e["date"])
            if top is None or e["e1rm"] > top["e1rm"]:
                per_date[e["date"]] = {
                    "e1rm": e["e1rm"], "reps": e["reps"], "kg": e["kg"],
                }
        dates_desc = sorted(per_date.keys(), reverse=True)
        if not dates_desc:
            continue
        current = per_date[dates_desc[0]]["e1rm"]
        prev = per_date[dates_desc[1]]["e1rm"] if len(dates_desc) >= 2 else None
        best = max(d["e1rm"] for d in per_date.values())

        # Slope is computed over the last 6 sessions for stability, even
        # though the emitted history is capped at E1RM_HISTORY_LIMIT.
        slope_dates = dates_desc[:6]
        history_full = [
            {
                "date":         d,
                "e1rm_kg":      round(per_date[d]["e1rm"], 1),
                "top_set_reps": per_date[d]["reps"],
                "top_set_kg":   per_date[d]["kg"],
            }
            for d in slope_dates
        ]
        history = history_full[:E1RM_HISTORY_LIMIT]

        # OLS slope (kg per 28 days) over the last 6 sessions. Use
        # ``history_full``, not the emitted ``history`` — the trim is
        # cosmetic for the JSON output, but the trend should still see all
        # six sessions to stay stable.
        slope = None
        if len(history_full) >= 3:
            pts: list[tuple[date, float]] = []
            for h in history_full:
                try:
                    pts.append((datetime.strptime(h["date"], "%Y-%m-%d").date(), h["e1rm_kg"]))
                except ValueError:
                    continue
            if len(pts) >= 3:
                pts.sort(key=lambda p: p[0])
                base = pts[0][0]
                xs = [(p[0] - base).days for p in pts]
                ys = [p[1] for p in pts]
                n = len(xs)
                mx = sum(xs) / n
                my = sum(ys) / n
                num = sum((xs[i] - mx) * (ys[i] - my) for i in range(n))
                den = sum((xs[i] - mx) ** 2 for i in range(n))
                if den > 0:
                    slope = round((num / den) * 28.0, 2)

        # Confidence from rep ranges of the last 3 sessions' top sets.
        # Epley is calibrated best for low-rep sets; 12+ rep top sets
        # give a noisy projection. Pulled from the un-capped history so
        # confidence is consistent regardless of the emitted limit.
        recent_reps = [h["top_set_reps"] for h in history_full[:3]]
        if len(recent_reps) < 2:
            confidence = "low"
        elif any(r >= 13 for r in recent_reps):
            confidence = "low"
        elif all(3 <= r <= 8 for r in recent_reps):
            confidence = "high"
        else:
            confidence = "medium"

        # Stalled: walk back through consecutive sessions while the
        # e1RM swing is within ±0.5kg. Break on the first deload that
        # falls inside (or at either end of) the gap between two
        # consecutive sessions — a deliberate volume cut isn't a stall.
        stalled = 0
        for i in range(len(dates_desc) - 1):
            this_date = dates_desc[i]
            prev_date = dates_desc[i + 1]
            crossed_deload = any(
                prev_date <= d <= this_date for d in deload_set
            )
            if crossed_deload:
                break
            this_e = per_date[this_date]["e1rm"]
            prev_e = per_date[prev_date]["e1rm"]
            if abs(this_e - prev_e) <= 0.5:
                stalled += 1
            else:
                break

        # Drop e1rm_history entirely unless explicitly opted in. The
        # summary fields (current/prev/best, slope, confidence,
        # stalled_sessions) cover every coaching decision; the per-session
        # history is debug-only and added ~10 KB to the default output.
        emit_history = include_history and not (confidence == "low" and slope is None)

        out[canonical_name[key]] = {
            "current_e1rm_kg":  round(current, 1),
            "prev_e1rm_kg":     round(prev, 1) if prev is not None else None,
            "best_e1rm_kg":     round(best, 1),
            "last_date":        dates_desc[0],
            "delta_vs_prev_kg": (round(current - prev, 1) if prev is not None else None),
            "e1rm_history":     history if emit_history else None,
            "slope_kg_per_4w":  slope,
            "confidence":       confidence,
            "stalled_sessions": stalled,
        }
    return out


def stale_exercises(
    rows: list[dict], db: dict[str, dict], today_d: date, threshold_days: int
) -> list[dict]:
    """Exercises whose last appearance is ≥ ``threshold_days`` ago.

    Warmup-section exercises are excluded — those cycle on and off by
    design. Useful for spotting movements that were tried once or twice and
    dropped; the coach can decide whether to retire or reintroduce them.
    """
    last_seen: dict[str, str] = {}
    sessions_count: dict[str, set[str]] = defaultdict(set)
    canonical: dict[str, str] = {}
    for r in rows:
        if not _is_working_set(r):
            continue
        key = r["exercise"].lower()
        entry = db.get(key)
        if entry and (entry.get("is_warmup") or entry.get("is_cardio")):
            continue
        canonical.setdefault(key, r["exercise"])
        if r["date"] > last_seen.get(key, ""):
            last_seen[key] = r["date"]
        sessions_count[key].add(r["date"])

    out = []
    for key, last in last_seen.items():
        try:
            d = datetime.strptime(last, "%Y-%m-%d").date()
        except ValueError:
            continue
        days = (today_d - d).days
        if days < threshold_days:
            continue
        out.append({
            "exercise":        canonical[key],
            "last_date":       last,
            "weeks_since":     round(days / 7.0, 1),
            "sessions_logged": len(sessions_count[key]),
        })
    out.sort(key=lambda e: e["weeks_since"], reverse=True)
    return out


# ============================================================================
# Derived metrics — turn raw session/health data into actionable signals.
# ============================================================================
#
# Everything below is pre-computed in Python so the coach LLM consumes
# decisions, not raw arrays. Each function returns ``None`` (or the
# closest empty form) when the source data is too sparse.

# Karvonen / HR-zone definitions (% of HRR — heart rate reserve).
HR_ZONES_PCT = [
    ("z1", 0.50, 0.60),
    ("z2", 0.60, 0.70),
    ("z3", 0.70, 0.80),
    ("z4", 0.80, 0.90),
    ("z5", 0.90, 1.00),
]


def _percentile(values: list[float], pct: float) -> float | None:
    """Return the ``pct`` percentile (0-100) of a numeric list, linear-interp."""
    if not values:
        return None
    s = sorted(values)
    if len(s) == 1:
        return s[0]
    k = (pct / 100.0) * (len(s) - 1)
    lo = int(k)
    hi = min(lo + 1, len(s) - 1)
    frac = k - lo
    return s[lo] * (1 - frac) + s[hi] * frac


def _age_from_birthday(birthday_str: str | None, today_d: date) -> int | None:
    """Return integer age in years from a YYYY-MM-DD birthday + today, or None."""
    if not birthday_str:
        return None
    try:
        bd = datetime.strptime(birthday_str, "%Y-%m-%d").date()
    except ValueError:
        return None
    age = today_d.year - bd.year - ((today_d.month, today_d.day) < (bd.month, bd.day))
    if age < 5 or age > 100:
        return None
    return age


def estimate_max_hr(workout_sessions_all: list[dict],
                    today_d: date,
                    profile: dict | None = None,
                    fallback_age: int = 30) -> float | None:
    """Estimate the user's peak HR.

    Priority:
    1. Highest observed ``max_hr`` across all Apple workouts — Apple's
       per-workout max is conservative (it reflects what the watch saw,
       not theoretical max), so the absolute observed peak is a sound
       lower-bound estimate of true max.
    2. Tanaka formula 208 − 0.7×age, with age computed dynamically from
       ``profile.birthday`` and ``today_d``. Falls back to ``fallback_age``
       (30) when birthday is missing or malformed.

    Returns ``None`` only when neither path can produce a number.
    """
    observed = [
        s.get("max_hr") for s in (workout_sessions_all or [])
        if s.get("max_hr") and s.get("max_hr") >= 140
    ]
    if observed:
        return float(max(observed))
    age = _age_from_birthday((profile or {}).get("birthday"), today_d)
    if age is None:
        age = fallback_age
    return float(round(208 - 0.7 * age, 1))


def health_metrics_weekly(health_all: list[dict],
                          today_d: date, weeks: int = 4) -> list[dict]:
    """Per-week aggregates of Health Metrics for the last N weeks.

    Replaces the 8 KB raw daily dump with a compact 4-week snapshot. Each
    entry is the mean of available daily values that landed in that ISO
    week (Mon-Sun). Only fields with at least one value in the window are
    emitted; sources that structurally can't provide a metric (HL slim
    schema) yield None for that key, which ``_compact`` strips.
    """
    if not health_all:
        return []
    cutoff = today_d - timedelta(days=weeks * 7)
    keys = [
        "vo2max", "resting_hr", "hrv_sdnn", "walking_hr",
        "hr_recovery_1min", "sleep_total_h", "sleep_deep_h", "sleep_rem_h",
        "resp_rate", "wrist_temp_c", "exercise_min",
    ]
    by_week: dict[tuple[int, int], dict[str, list[float]]] = {}
    for e in health_all:
        try:
            d = datetime.strptime(e["date"], "%Y-%m-%d").date()
        except (ValueError, KeyError):
            continue
        if d < cutoff:
            continue
        iso = d.isocalendar()
        wk = (iso.year, iso.week)
        bucket = by_week.setdefault(wk, {})
        for k in keys:
            v = e.get(k)
            if v is None:
                continue
            bucket.setdefault(k, []).append(float(v))

    out: list[dict] = []
    for wk in sorted(by_week.keys()):
        # Monday of the ISO week — readable anchor for the LLM.
        monday = datetime.fromisocalendar(wk[0], wk[1], 1).date()
        entry: dict = {"week_start": monday.strftime("%Y-%m-%d"),
                       "n_days": max(len(v) for v in by_week[wk].values())}
        for k in keys:
            vals = by_week[wk].get(k)
            if not vals:
                entry[k] = None
                continue
            entry[k] = round(sum(vals) / len(vals), 2)
        out.append(entry)
    return out


def recovery_score(health_all: list[dict], today_d: date,
                   capabilities: dict) -> dict:
    """Compute a 0-10 recovery score from HRV, RHR, sleep, wrist temp,
    HR Recovery 1-min, and VO2max trend.

    Each signal contributes a clamped delta to a baseline of 5; the final
    score is clamped to [0, 10]. Drivers list shows which signals
    dominated so the coach can explain *why*.
    Returns ``{"score": float, "drivers": list[dict], "confidence": str}``.
    """
    drivers: list[dict] = []
    score = 5.0
    sample_count = 0

    # HRV: reading vs personal 60d baseline. ±10% baseline = ±2 score swing.
    if capabilities.get("hrv"):
        recent = _values_in_window(health_all, "hrv_sdnn", today_d, 7)
        baseline = baseline_60d(health_all, "hrv_sdnn", today_d)
        if recent and baseline and baseline > 0:
            recent_avg = sum(recent) / len(recent)
            delta = (recent_avg - baseline) / baseline
            contrib = max(-2.0, min(2.0, delta / 0.05))  # ±5% baseline → ±1
            score += contrib
            sample_count += 1
            drivers.append({
                "metric":     "hrv_sdnn",
                "recent_avg": round(recent_avg, 1),
                "baseline":   round(baseline, 1),
                "delta_pct":  round(delta * 100, 1),
                "contrib":    round(contrib, 2),
            })

    # Resting HR: lower is better. ±5 bpm vs 28d typical → ±2 swing.
    rhr_recent = _values_in_window(health_all, "resting_hr", today_d, 7)
    rhr_typical = _values_in_window(health_all, "resting_hr", today_d, 28)
    if rhr_recent and rhr_typical:
        recent_avg = sum(rhr_recent) / len(rhr_recent)
        typical_avg = sum(rhr_typical) / len(rhr_typical)
        delta = recent_avg - typical_avg
        contrib = max(-2.0, min(2.0, -delta / 2.5))  # +2.5 bpm → -1
        score += contrib
        sample_count += 1
        drivers.append({
            "metric":      "resting_hr",
            "recent_avg":  round(recent_avg, 1),
            "typical_avg": round(typical_avg, 1),
            "delta_bpm":   round(delta, 1),
            "contrib":     round(contrib, 2),
        })

    # Sleep: 7h target. ±1h → ±1 swing (clamped ±2).
    sleep = _values_in_window(health_all, "sleep_total_h", today_d, 7)
    if sleep:
        recent_avg = sum(sleep) / len(sleep)
        delta = recent_avg - 7.0
        contrib = max(-2.0, min(2.0, delta))
        score += contrib
        sample_count += 1
        drivers.append({
            "metric":     "sleep_total_h",
            "recent_avg": round(recent_avg, 2),
            "target":     7.0,
            "delta_h":    round(delta, 2),
            "contrib":    round(contrib, 2),
        })

    # Wrist temp: deviation from 60d baseline. >+0.3°C is a stress/illness
    # signal; weight modestly (max ±1.5).
    if capabilities.get("wrist_temp"):
        wt_recent = _values_in_window(health_all, "wrist_temp_c", today_d, 3)
        wt_base = baseline_60d(health_all, "wrist_temp_c", today_d)
        if wt_recent and wt_base:
            recent_avg = sum(wt_recent) / len(wt_recent)
            delta = recent_avg - wt_base
            contrib = max(-1.5, min(1.5, -delta / 0.2))
            score += contrib
            sample_count += 1
            drivers.append({
                "metric":     "wrist_temp_c",
                "recent_avg": round(recent_avg, 2),
                "baseline":   round(wt_base, 2),
                "delta_c":    round(delta, 2),
                "contrib":    round(contrib, 2),
            })

    # HR Recovery 1-min (count/min HR drop after exercise). Higher is
    # better — parasympathetic re-activation. Weight ±0.75. Compare
    # recent (5d) vs 28d typical; ±5 bpm = ±0.75.
    hrr_recent = _values_in_window(health_all, "hr_recovery_1min", today_d, 5)
    hrr_typical = _values_in_window(health_all, "hr_recovery_1min", today_d, 28)
    if hrr_recent and hrr_typical:
        recent_avg = sum(hrr_recent) / len(hrr_recent)
        typical_avg = sum(hrr_typical) / len(hrr_typical)
        delta = recent_avg - typical_avg
        contrib = max(-0.75, min(0.75, delta / 6.7))  # +5 bpm → +0.75
        score += contrib
        sample_count += 1
        drivers.append({
            "metric":      "hr_recovery_1min",
            "recent_avg":  round(recent_avg, 1),
            "typical_avg": round(typical_avg, 1),
            "delta_bpm":   round(delta, 1),
            "contrib":     round(contrib, 2),
        })

    # VO2max trend per 4 weeks. Slow-moving signal — folding it in gives
    # credit for fitness improvements / penalises drift. Weight ±0.75;
    # ±2 ml/kg/min over 4w → ±0.75.
    vo2_slope = metric_trend_per_4w(health_all, "vo2max")
    if vo2_slope is not None:
        contrib = max(-0.75, min(0.75, vo2_slope / 2.7))
        score += contrib
        sample_count += 1
        drivers.append({
            "metric":      "vo2max_trend_per_4w",
            "slope":       round(vo2_slope, 2),
            "contrib":     round(contrib, 2),
        })

    score = max(0.0, min(10.0, score))
    confidence = "high" if sample_count >= 3 else ("medium" if sample_count == 2 else "low")
    return {
        "score":      round(score, 1),
        "confidence": confidence,
        "drivers":    drivers,
    }


def _trimp(duration_min: float, avg_hr: float,
           rest_hr: float, max_hr: float) -> float:
    """Banister TRIMP. ``duration_min × HRr × 0.64 × e^(1.92×HRr)`` (men).

    Only positive when HR is above resting. Uses HRR (heart rate
    reserve) normalisation so the same TRIMP score means the same
    relative effort across users.
    """
    import math
    if not duration_min or not avg_hr or not max_hr or not rest_hr:
        return 0.0
    if max_hr <= rest_hr:
        return 0.0
    hrr = (avg_hr - rest_hr) / (max_hr - rest_hr)
    if hrr <= 0:
        return 0.0
    hrr = min(hrr, 1.0)
    return round(duration_min * hrr * 0.64 * math.exp(1.92 * hrr), 1)


def trimp_per_session(monthly_sessions: list[dict],
                      max_hr: float | None,
                      rest_hr: float | None) -> list[dict]:
    """Compute TRIMP for every session that has both avg_hr and duration.

    Returns one entry per session with ``date, kind, trimp, intensity_pct``
    (HRR percent), plus a ``load_band`` classification: light <50,
    moderate 50-100, hard 100-150, red-line >150.
    """
    if not max_hr or not rest_hr or max_hr <= rest_hr:
        return []
    out: list[dict] = []
    for s in monthly_sessions:
        avg_hr = s.get("avg_hr")
        dur = s.get("duration_min")
        if not avg_hr or not dur:
            continue
        try:
            avg_hr_f = float(avg_hr)
            dur_f = float(dur)
        except (TypeError, ValueError):
            continue
        trimp = _trimp(dur_f, avg_hr_f, rest_hr, max_hr)
        if trimp == 0:
            continue
        hrr_pct = (avg_hr_f - rest_hr) / (max_hr - rest_hr)
        hrr_pct = max(0.0, min(1.0, hrr_pct))
        if trimp < 50:
            band = "light"
        elif trimp < 100:
            band = "moderate"
        elif trimp < 150:
            band = "hard"
        else:
            band = "red-line"
        out.append({
            "date":          s["date"],
            "kind":          s.get("session_kind", "other"),
            "trimp":         trimp,
            "intensity_pct": round(hrr_pct * 100, 1),
            "load_band":     band,
        })
    out.sort(key=lambda e: e["date"])
    return out


def training_load_summary(trimps: list[dict], today_d: date) -> dict:
    """CTL (chronic, 42d EWMA), ATL (acute, 7d EWMA), TSB (form = CTL−ATL).

    Standard TrainingPeaks formulas. CTL ≈ fitness, ATL ≈ fatigue, TSB
    positive = peaked, negative = under load. Computed by walking each
    day from the earliest TRIMP to today and decaying yesterday's value.
    Returns the values *as of today_d*.
    """
    if not trimps:
        return {"ctl": None, "atl": None, "tsb": None, "trend_7d": None}
    # Convert to a date→trimp dict (sum if multiple sessions same day).
    by_date: dict[date, float] = {}
    for t in trimps:
        try:
            d = datetime.strptime(t["date"], "%Y-%m-%d").date()
        except ValueError:
            continue
        by_date[d] = by_date.get(d, 0.0) + t["trimp"]
    if not by_date:
        return {"ctl": None, "atl": None, "tsb": None, "trend_7d": None}
    start = min(by_date.keys())
    ctl_alpha = 1.0 / 42.0  # ~time constant 42d
    atl_alpha = 1.0 / 7.0
    ctl = atl = 0.0
    history: list[tuple[date, float, float]] = []
    cur = start
    while cur <= today_d:
        load = by_date.get(cur, 0.0)
        ctl = ctl + ctl_alpha * (load - ctl)
        atl = atl + atl_alpha * (load - atl)
        history.append((cur, ctl, atl))
        cur += timedelta(days=1)
    today_ctl, today_atl = ctl, atl
    week_ago_ctl = next(
        (h[1] for h in reversed(history)
         if h[0] <= today_d - timedelta(days=7)),
        today_ctl
    )
    return {
        "ctl":      round(today_ctl, 1),
        "atl":      round(today_atl, 1),
        "tsb":      round(today_ctl - today_atl, 1),
        "trend_7d": round(today_ctl - week_ago_ctl, 1),
    }


def hr_at_volume_divergence(rows: list[dict],
                             monthly_sessions: list[dict],
                             db: dict, today_d: date,
                             window_weeks: int = 8) -> dict:
    """Per-muscle-group HR-creep signal at constant volume.

    For each muscle group, regress ``session_avg_hr`` against time over
    the last ``window_weeks`` weeks of strength sessions, weighting by
    that session's volume into the muscle. Positive slope (HR rising at
    same volume) suggests fatigue; negative slope is improving
    conditioning. Returns ``{muscle: {slope_bpm_per_4w, n_sessions, hint}}``.
    """
    if not monthly_sessions:
        return {}
    cutoff = today_d - timedelta(days=window_weeks * 7)
    # Build date → strength session avg_hr lookup.
    strength_hr: dict[str, float] = {}
    for s in monthly_sessions:
        if s.get("session_kind") != "strength":
            continue
        if s.get("avg_hr") in (None, 0):
            continue
        try:
            d = datetime.strptime(s["date"], "%Y-%m-%d").date()
        except ValueError:
            continue
        if d < cutoff:
            continue
        strength_hr[s["date"]] = float(s["avg_hr"])
    if len(strength_hr) < 4:
        return {}

    # Roll up rows by (date, muscle) → volume.
    per_date_muscle: dict[tuple[str, str], float] = {}
    for r in rows:
        if not _is_working_set(r):
            continue
        if r["date"] not in strength_hr:
            continue
        muscles = db.get(r["exercise"].lower())
        if not muscles:
            continue
        primary = muscles.get("primary") if isinstance(muscles, dict) else None
        if not primary:
            continue
        vol = (r.get("volume") or 0)
        per_date_muscle[(r["date"], primary)] = (
            per_date_muscle.get((r["date"], primary), 0.0) + vol
        )

    by_muscle: dict[str, list[tuple[date, float, float]]] = {}
    for (d_str, muscle), vol in per_date_muscle.items():
        if vol <= 0:
            continue
        try:
            d = datetime.strptime(d_str, "%Y-%m-%d").date()
        except ValueError:
            continue
        by_muscle.setdefault(muscle, []).append((d, strength_hr[d_str], vol))

    out: dict[str, dict] = {}
    for muscle, points in by_muscle.items():
        # Require at least 6 sessions before a slope is published —
        # smaller samples have too much variance for the ±5 bpm/4w
        # threshold to mean anything.
        if len(points) < 6:
            continue
        points.sort(key=lambda p: p[0])
        base = points[0][0]
        xs = [(p[0] - base).days for p in points]
        ys = [p[1] for p in points]
        ws = [p[2] for p in points]
        sum_w = sum(ws)
        if sum_w <= 0:
            continue
        mx = sum(xs[i] * ws[i] for i in range(len(xs))) / sum_w
        my = sum(ys[i] * ws[i] for i in range(len(ys))) / sum_w
        num = sum(ws[i] * (xs[i] - mx) * (ys[i] - my) for i in range(len(xs)))
        den = sum(ws[i] * (xs[i] - mx) ** 2 for i in range(len(xs)))
        if den <= 0:
            continue
        slope_per_day = num / den
        slope_per_4w = slope_per_day * 28
        # ±5 bpm/4w is the magnitude that's clearly above noise floor for
        # a 6-12 session window. Below that, call it stable to avoid
        # crying wolf on every minor drift.
        if slope_per_4w >= 5:
            hint = "rising HR at constant volume — fatigue or under-recovery"
        elif slope_per_4w <= -5:
            hint = "falling HR at constant volume — improving conditioning"
        else:
            hint = "stable"
        out[muscle] = {
            "slope_bpm_per_4w": round(slope_per_4w, 2),
            "n_sessions":       len(points),
            "hint":             hint,
        }
    return out


def cardio_hr_zones(monthly_sessions: list[dict],
                    today_d: date,
                    max_hr: float | None,
                    rest_hr: float | None,
                    window_days: int = 28) -> dict:
    """Polarized-vs-pyramidal HR distribution across cardio sessions.

    Without per-second HR data we can't compute true time-in-zone, so we
    place each session entirely in the zone its avg_hr falls into.
    Coarse but useful for trend (has the user been doing too much Z3
    grey-zone work?). Returns ``{z1..z5: minutes, total_minutes,
    polarized_pct, pyramidal_pct, threshold_pct}``.
    """
    if not max_hr or not rest_hr or max_hr <= rest_hr:
        return {}
    cutoff = today_d - timedelta(days=window_days)
    zone_min: dict[str, float] = {z[0]: 0.0 for z in HR_ZONES_PCT}
    total = 0.0
    for s in monthly_sessions:
        if s.get("session_kind") != "cardio":
            continue
        try:
            d = datetime.strptime(s["date"], "%Y-%m-%d").date()
        except (ValueError, KeyError):
            continue
        if d < cutoff:
            continue
        avg_hr = s.get("avg_hr")
        dur = s.get("duration_min")
        if not avg_hr or not dur:
            continue
        hrr = (float(avg_hr) - rest_hr) / (max_hr - rest_hr)
        hrr = max(0.0, min(1.0, hrr))
        for label, lo, hi in HR_ZONES_PCT:
            if hrr < hi or label == "z5":
                zone_min[label] += float(dur)
                total += float(dur)
                break
    if total <= 0:
        return {}
    z1 = round(zone_min["z1"], 1)
    z2 = round(zone_min["z2"], 1)
    z3 = round(zone_min["z3"], 1)
    z4 = round(zone_min["z4"], 1)
    z5 = round(zone_min["z5"], 1)
    return {
        "window_days":    window_days,
        "total_minutes":  round(total, 1),
        "z1": z1, "z2": z2, "z3": z3, "z4": z4, "z5": z5,
        "z2_pct": round((z2 / total) * 100, 1),
        "z3_pct": round((z3 / total) * 100, 1),
        "z4_z5_pct": round(((z4 + z5) / total) * 100, 1),
    }


def auto_deload_candidates(monthly_sessions: list[dict],
                           deloads_logged: list[str],
                           today_d: date,
                           window_weeks: int = 8) -> list[str]:
    """Detect strength-session weeks where volume + HR both dropped enough
    to look like a deload that the user didn't mark.

    Heuristic per week:
    - Median session volume ≤ 0.65 × prior 4-week median.
    - AND median session avg_hr ≤ prior_4wk_median - 8 bpm.
    - AND not already in ``deloads_logged``.

    Conservative — designed to surface candidates the user likely forgot
    to flag, not to second-guess intent.
    """
    cutoff = today_d - timedelta(days=window_weeks * 7)
    strength = []
    for s in monthly_sessions:
        if s.get("session_kind") != "strength":
            continue
        try:
            d = datetime.strptime(s["date"], "%Y-%m-%d").date()
        except (ValueError, KeyError):
            continue
        if d < cutoff:
            continue
        vol = s.get("volume")
        hr = s.get("avg_hr")
        if not vol or not hr:
            continue
        strength.append((d, float(vol), float(hr)))
    if len(strength) < 6:
        return []
    strength.sort(key=lambda p: p[0])

    def median(xs):
        s = sorted(xs)
        n = len(s)
        return s[n // 2] if n % 2 else (s[n // 2 - 1] + s[n // 2]) / 2

    candidates: list[str] = []
    for i, (d, vol, hr) in enumerate(strength):
        prior = [p for p in strength[:i]
                 if (d - p[0]).days <= 28 and (d - p[0]).days > 0]
        if len(prior) < 4:
            continue
        prior_vol = median([p[1] for p in prior])
        prior_hr = median([p[2] for p in prior])
        if prior_vol <= 0:
            continue
        date_str_form = d.strftime("%Y-%m-%d")
        if date_str_form in deloads_logged:
            continue
        if vol <= 0.65 * prior_vol and hr <= prior_hr - 8:
            candidates.append(date_str_form)
    return candidates


# ---------- main ----------
def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("tracker", type=Path)
    ap.add_argument("--months", type=int, default=3,
                    help="How many months back to load from monthly sheets. The data is used internally for "
                         "all roll-ups regardless of --include-rows.")
    ap.add_argument("--today", default=None, help="Override today's date (YYYY-MM-DD) for testing")
    ap.add_argument("--include-rows", action="store_true",
                    help="Include the flat `rows` array in the JSON. Off by default — the coach already exposes "
                         "pre-aggregated `monthly_sessions`, `progression_summary`, `weekly_volume_per_muscle`, "
                         "and `estimated_1rm`. Pass this only for debug deep-dives.")
    ap.add_argument("--include-1rm-history", action="store_true",
                    help="Include the per-exercise `e1rm_history` list (last 3 sessions). Off by default — "
                         "`current_e1rm_kg`, `slope_kg_per_4w`, `confidence`, and `stalled_sessions` cover the "
                         "coaching decision; the history is debug-only.")
    ap.add_argument("--include-daily-health", action="store_true",
                    help="Include the raw daily `health_metrics_recent` (~30 rows × 13 fields). Off by default — "
                         "the coach reads weekly aggregates from `health_metrics_weekly` instead.")
    ap.add_argument("--pretty", action="store_true",
                    help="Pretty-print the JSON (indent=2). Off by default — compact form saves ~20%% of "
                         "tokens for the LLM consumer. Use for human inspection.")
    args = ap.parse_args()

    if not args.tracker.exists():
        print(f"ERROR: tracker not found: {args.tracker}", file=sys.stderr)
        return 1

    today_d = (
        datetime.strptime(args.today, "%Y-%m-%d").date() if args.today else date.today()
    )

    wb = openpyxl.load_workbook(args.tracker, data_only=True, read_only=True)
    rows, session_totals, session_summaries = extract_rows(wb, args.months, today_d)
    deloads = find_deloads(wb)

    profile = read_profile(wb)
    data_source = profile.get("source") or DEFAULT_DATA_SOURCE
    capabilities = SOURCE_CAPABILITIES.get(data_source, SOURCE_CAPABILITIES[DEFAULT_DATA_SOURCE])

    last_session = max((r["date"] for r in rows), default=None)
    days_since = None
    if last_session:
        d = datetime.strptime(last_session, "%Y-%m-%d").date()
        days_since = (today_d - d).days

    weeks_since_deload = None
    if deloads:
        d = datetime.strptime(deloads[-1], "%Y-%m-%d").date()
        weeks_since_deload = round((today_d - d).days / 7.0, 1)

    health_all = read_health_metrics(wb)
    # Per-day rows go into health_metrics_recent. ``bodyweight_kg`` is dropped
    # because it duplicates the dedicated ``bodyweight_recent`` series — the
    # coach reads daily metrics for HRV / VO2max / sleep / wrist temp, not
    # weight. Keeping it here costs ~600 bytes for no signal.
    health_recent = [
        {k: v for k, v in entry.items() if k != "bodyweight_kg"}
        for entry in health_all[-30:]
    ]

    workout_sessions_all = read_workout_sessions(wb)
    sessions_28d = workout_sessions_in_window(workout_sessions_all, today_d, 28)

    # Strength session dates: any logged date with at least one working set
    # carrying weight × reps. Used to match Apple workouts back to training.
    strength_dates: set[str] = set()
    for r in rows:
        if not _is_working_set(r):
            continue
        if (r.get("kg") or 0) > 0 and (r.get("reps") or 0) > 0:
            strength_dates.add(r["date"])

    bw_all = read_bodyweight(wb)
    bw_recent = bw_all[-12:]
    bw_latest = (
        {"date": bw_all[-1]["date"], "kg": bw_all[-1]["kg"]}
        if bw_all else None
    )

    # Parse the exercises database once; it's read-only for this run.
    db_path = Path(__file__).resolve().parents[2] / "shared" / "exercises-database.md"
    db = load_exercises_db(db_path)

    unknown_set: set[str] = set()
    weekly_volume = weekly_volume_per_muscle(rows, db, today_d, 28, unknown_set)
    e1rm = estimated_1rm(rows, deloads,
                         include_history=args.include_1rm_history)
    stale_full = stale_exercises(rows, db, today_d, 28)
    # Cap stale_exercises to top 5 by weeks_since (already DESC-sorted by
    # the helper). Beyond 5 the coach rarely uses them in plan generation.
    stale = stale_full[:5]

    # Surface any logged exercise across the full loaded window that doesn't
    # match an entry in the database — not just the 28-day volume window.
    # Catches typos/rename drift (e.g. "Deadhang" vs "Dead Hang") that would
    # otherwise silently under-count volume and dodge rotation decisions.
    for r in rows:
        if not _is_working_set(r):
            continue
        if r["exercise"].lower() not in db:
            unknown_set.add(r["exercise"])

    # ---- Derived metrics ----
    monthly_sessions = build_monthly_sessions(
        rows, session_summaries,
        session_totals=session_totals,
        apple_sessions=workout_sessions_all,
    )
    max_hr = estimate_max_hr(workout_sessions_all, today_d, profile=profile)
    age_years = _age_from_birthday(profile.get("birthday"), today_d)
    rest_hr = _mean_or_none(_values_in_window(health_all, "resting_hr", today_d, 28))
    if rest_hr is None and capabilities.get("resting_hr_daily") is False:
        # HL fallback: typical adult RHR if the source can't supply it.
        rest_hr = 60.0

    recovery = recovery_score(health_all, today_d, capabilities)
    trimps = trimp_per_session(monthly_sessions, max_hr, rest_hr)
    training_load = training_load_summary(trimps, today_d)
    # Fold TRIMP load_band back onto each monthly_session for the LLM.
    trimp_by_date: dict[str, dict] = {t["date"]: t for t in trimps}
    for s in monthly_sessions:
        t = trimp_by_date.get(s.get("date"))
        if t:
            s["trimp"] = t["trimp"]
            s["load_band"] = t["load_band"]
            s["intensity_pct"] = t["intensity_pct"]

    hr_volume_div = hr_at_volume_divergence(rows, monthly_sessions, db, today_d)
    cardio_zones = cardio_hr_zones(monthly_sessions, today_d, max_hr, rest_hr)
    auto_deloads = auto_deload_candidates(monthly_sessions, deloads, today_d)
    weekly_health = health_metrics_weekly(health_all, today_d, weeks=4)

    out = {
        "today": today_d.strftime("%Y-%m-%d"),
        "data_source": data_source,
        "capabilities": capabilities,
        "auto_cardio_enabled": bool(profile.get("auto_cardio")),
        # ---- Strength + cardio sessions (canonical session-level view) ----
        "monthly_sessions": monthly_sessions,
        "weekly_volume_per_muscle": weekly_volume,
        "estimated_1rm": e1rm,
        "progression_summary": progression_summary(rows),
        "stale_exercises": stale,
        "unknown_exercises": sorted(unknown_set),
        "deloads": deloads,
        "auto_deload_candidates": auto_deloads,
        # ---- Cardio rollup ----
        "cardio_last_28d": cardio_last_28d(rows, today_d),
        "cardio_hr_zones_28d": cardio_zones,
        # ---- Recovery + training load (Python-derived, not raw metrics) ----
        "recovery": recovery,
        "training_load": training_load,
        "hr_at_volume_divergence": hr_volume_div,
        "age_years": age_years,
        "estimated_max_hr": max_hr,
        "estimated_rest_hr": round(rest_hr, 1) if rest_hr else None,
        # ---- Bodyweight ----
        "bodyweight_latest": bw_latest,
        "bodyweight_trend_kg_per_week": bodyweight_trend_kg_per_week(bw_all),
        # ---- Apple Health weekly aggregates (raw daily behind a flag) ----
        "health_metrics_weekly": weekly_health,
        "health_metrics_recent": health_recent if args.include_daily_health else None,
        "vo2max_latest": latest_metric(health_all, "vo2max"),
        "vo2max_trend_per_4w": metric_trend_per_4w(health_all, "vo2max"),
        # ---- Debug deep-dive: flat per-set list (--include-rows). ----
        "rows": rows if args.include_rows else None,
    }
    if args.pretty:
        json.dump(_compact(out), sys.stdout, ensure_ascii=False, indent=2)
    else:
        # Compact form: no whitespace between separators. Saves ~20% of
        # bytes vs indent=2 for an LLM consumer that doesn't render the
        # whitespace anyway.
        json.dump(_compact(out), sys.stdout, ensure_ascii=False,
                  separators=(",", ":"))
    return 0


if __name__ == "__main__":
    sys.exit(main())
